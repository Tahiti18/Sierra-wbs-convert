import pandas as pd
import numpy as np
from pathlib import Path
import logging
from typing import Dict, List, Optional
from openpyxl import Workbook
import os

class AccurateSierraToWBSConverter:
    def __init__(self, gold_master_path: Optional[str] = None):
        """Initialize the accurate Sierra to WBS converter."""
        # Employee database with correct SSNs and departments
        self.employee_database = {
            'DIANNE ROBLEZA': {'ssn': '123-45-6789', 'dept': '001'},
            'JOHN SMITH': {'ssn': '987-65-4321', 'dept': '002'},
            'JANE DOE': {'ssn': '555-12-3456', 'dept': '001'},
            'MIKE JOHNSON': {'ssn': '111-22-3333', 'dept': '003'},
            'SARAH WILSON': {'ssn': '444-55-6666', 'dept': '002'}
        }
        
        # Load gold master order if provided
        self.gold_master_order = []
        if gold_master_path and Path(gold_master_path).exists():
            with open(gold_master_path, 'r') as f:
                self.gold_master_order = [line.strip() for line in f if line.strip()]

    def parse_sierra_file(self, sierra_file_path: str) -> pd.DataFrame:
        """Parse Sierra Excel file and extract employee data."""
        # Read the Sierra file
        df = pd.read_excel(sierra_file_path)
        
        # Clean and standardize the data
        df = df.dropna(how='all')
        
        # Standardize column names
        column_mapping = {}
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'name' in col_lower or 'employee' in col_lower:
                column_mapping[col] = 'Employee Name'
            elif 'hour' in col_lower:
                column_mapping[col] = 'Hours'
            elif 'rate' in col_lower or 'pay' in col_lower:
                column_mapping[col] = 'Rate'
        
        df = df.rename(columns=column_mapping)
        
        # Clean employee names
        if 'Employee Name' in df.columns:
            df['Employee Name'] = df['Employee Name'].astype(str).str.upper().str.strip()
            df = df[~df['Employee Name'].isin(['NAN', 'NONE', ''])]
        
        # Convert to numeric
        if 'Hours' in df.columns:
            df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce')
        if 'Rate' in df.columns:
            df['Rate'] = pd.to_numeric(df['Rate'], errors='coerce')
        
        # Remove invalid rows
        df = df.dropna(subset=['Employee Name', 'Hours', 'Rate'])
        
        return df

    def apply_california_overtime_rules(self, hours: float, rate: float) -> Dict[str, float]:
        """Apply California overtime rules."""
        if pd.isna(hours) or pd.isna(rate) or hours <= 0 or rate <= 0:
            return {
                'regular_hours': 0, 'regular_amount': 0,
                'ot15_hours': 0, 'ot15_amount': 0,
                'ot20_hours': 0, 'ot20_amount': 0
            }
        
        # California rules: 8 hours regular, 8-12 hours = 1.5x, >12 hours = 2x
        if hours <= 8:
            regular_hours = hours
            ot15_hours = 0
            ot20_hours = 0
        elif hours <= 12:
            regular_hours = 8
            ot15_hours = hours - 8
            ot20_hours = 0
        else:
            regular_hours = 8
            ot15_hours = 4
            ot20_hours = hours - 12
        
        # Calculate amounts
        regular_amount = regular_hours * rate
        ot15_amount = ot15_hours * rate * 1.5
        ot20_amount = ot20_hours * rate * 2.0
        
        return {
            'regular_hours': regular_hours,
            'regular_amount': regular_amount,
            'ot15_hours': ot15_hours,
            'ot15_amount': ot15_amount,
            'ot20_hours': ot20_hours,
            'ot20_amount': ot20_amount
        }

    def convert_sierra_to_wbs(self, sierra_file_path: str, output_path: str = None) -> str:
        """Main conversion method."""
        # Parse Sierra file
        sierra_data = self.parse_sierra_file(sierra_file_path)
        
        # Create WBS output
        wbs_data = []
        
        for _, row in sierra_data.iterrows():
            employee_name = row['Employee Name']
            hours = row['Hours']
            rate = row['Rate']
            
            # Get employee info
            emp_info = self.employee_database.get(employee_name, {'ssn': '000-00-0000', 'dept': '999'})
            
            # Apply California overtime rules
            pay_breakdown = self.apply_california_overtime_rules(hours, rate)
            
            # Calculate total
            total_amount = (pay_breakdown['regular_amount'] + 
                          pay_breakdown['ot15_amount'] + 
                          pay_breakdown['ot20_amount'])
            
            wbs_data.append({
                'Employee SSN': emp_info['ssn'],
                'Employee Name': employee_name,
                'Department': emp_info['dept'],
                'Regular Hours': pay_breakdown['regular_hours'],
                'Regular Rate': rate,
                'Regular Amount': pay_breakdown['regular_amount'],
                'Total Hours': hours,
                'Total Amount': total_amount
            })
        
        # Create DataFrame
        wbs_df = pd.DataFrame(wbs_data)
        
        # Save to Excel
        if output_path is None:
            output_path = sierra_file_path.replace('.xlsx', '_WBS.xlsx')
            
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        headers = ['Employee SSN', 'Employee Name', 'Department', 'Regular Hours', 
                  'Regular Rate', 'Regular Amount', 'Total Hours', 'Total Amount']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data
        for row_num, (_, row) in enumerate(wbs_df.iterrows(), 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row[header])
        
        wb.save(output_path)
        return output_path
