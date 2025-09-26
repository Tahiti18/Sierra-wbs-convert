#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook

def compare_final_accuracy():
    """Compare our final conversion against actual WBS"""
    
    print("FINAL ACCURACY COMPARISON")
    print("=" * 80)
    
    # Load actual WBS
    actual_wb = load_workbook("/home/user/webapp/WBS_Payroll_9_12_25_for_Marwan_2.xlsx", data_only=True)
    actual_ws = actual_wb.active
    
    # Load our conversion
    our_wb = load_workbook("/home/user/webapp/test_final_conversion.xlsx", data_only=True)
    our_ws = our_wb.active
    
    print("COMPARISON - First 15 employees:")
    print("Format: Name | Actual (Rate, Hours, Total) | Ours (Rate, Hours, Total) | Match?")
    print("-" * 100)
    
    matches = 0
    total_employees = 0
    actual_total = 0
    our_total = 0
    
    for row_num in range(9, 24):  # Check first 15 employee rows
        # Actual WBS
        actual_name = actual_ws.cell(row=row_num, column=3).value
        actual_rate = actual_ws.cell(row=row_num, column=6).value
        actual_hours = actual_ws.cell(row=row_num, column=8).value
        actual_amount = actual_ws.cell(row=row_num, column=28).value
        
        # Our WBS
        our_name = our_ws.cell(row=row_num, column=3).value
        our_rate = our_ws.cell(row=row_num, column=6).value
        our_hours = our_ws.cell(row=row_num, column=8).value
        our_amount = our_ws.cell(row=row_num, column=28).value
        
        if actual_name:
            total_employees += 1
            
            # Format values for comparison
            def fmt_val(val):
                if val is None:
                    return "None"
                elif isinstance(val, (int, float)):
                    return f"{val}"
                else:
                    return str(val)
            
            actual_str = f"({fmt_val(actual_rate)}, {fmt_val(actual_hours)}, {fmt_val(actual_amount)})"
            our_str = f"({fmt_val(our_rate)}, {fmt_val(our_hours)}, {fmt_val(our_amount)})"
            
            # Check if amounts match (most important)
            amount_match = (actual_amount == our_amount) or (actual_amount is None and our_amount is None)
            
            if amount_match:
                matches += 1
                status = "✅"
            else:
                status = "❌"
            
            print(f"{actual_name:25} | {actual_str:20} | {our_str:20} | {status}")
            
            # Add to totals
            if actual_amount:
                actual_total += float(actual_amount)
            if our_amount:
                our_total += float(our_amount)
    
    print("-" * 100)
    print(f"SUMMARY:")
    print(f"  Employees checked: {total_employees}")
    print(f"  Exact matches: {matches}")
    print(f"  Accuracy: {matches/total_employees*100:.1f}%")
    print(f"  Actual total: ${actual_total:.2f}")
    print(f"  Our total: ${our_total:.2f}")
    print(f"  Difference: ${our_total - actual_total:.2f}")
    
    # Check specific discrepancies
    print(f"\n" + "=" * 80)
    print("DETAILED DISCREPANCIES:")
    
    discrepancies = []
    for row_num in range(9, 24):
        actual_name = actual_ws.cell(row=row_num, column=3).value
        actual_amount = actual_ws.cell(row=row_num, column=28).value
        our_amount = our_ws.cell(row=row_num, column=28).value
        
        if actual_name and actual_amount != our_amount:
            discrepancies.append({
                'name': actual_name,
                'actual': actual_amount,
                'ours': our_amount,
                'diff': (our_amount or 0) - (actual_amount or 0)
            })
    
    for disc in discrepancies:
        print(f"  {disc['name']}: Actual=${disc['actual']}, Ours=${disc['ours']}, Diff=${disc['diff']:.2f}")

if __name__ == "__main__":
    compare_final_accuracy()