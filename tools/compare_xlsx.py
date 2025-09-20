# tools/compare_xlsx.py
import openpyxl
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path

def cell_value_str(v):
    if v is None:
        return ""
    return str(v).strip()

def compare_files(a_path, b_path, out_report="compare_report.txt"):
    a_wb = openpyxl.load_workbook(a_path, data_only=True)
    b_wb = openpyxl.load_workbook(b_path, data_only=True)
    report = []
    sheets = set(a_wb.sheetnames) | set(b_wb.sheetnames)
    for s in sheets:
        report.append(f"=== Sheet: {s} ===")
        a_ws = a_wb[s] if s in a_wb.sheetnames else None
        b_ws = b_wb[s] if s in b_wb.sheetnames else None
        if a_ws is None:
            report.append(f" missing in A")
            continue
        if b_ws is None:
            report.append(f" missing in B")
            continue
        max_row = max(a_ws.max_row, b_ws.max_row)
        max_col = max(a_ws.max_column, b_ws.max_column)
        diff_count = 0
        for r in range(1, max_row+1):
            for c in range(1, max_col+1):
                a_val = cell_value_str(a_ws.cell(row=r, column=c).value)
                b_val = cell_value_str(b_ws.cell(row=r, column=c).value)
                if a_val != b_val:
                    diff_count += 1
                    report.append(f" R{r}C{c} ({get_column_letter(c)}{r}): A='{a_val}'  B='{b_val}'")
                    if diff_count > 200:
                        report.append(" ... more differences truncated ...")
                        break
            if diff_count > 200:
                break
        report.append(f" Sheet diffs: {diff_count}")
    Path(out_report).write_text("\n".join(report), encoding="utf-8")
    print("Report written to", out_report)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python tools/compare_xlsx.py generated.xlsx gold.xlsx")
        sys.exit(1)
    compare_files(sys.argv[1], sys.argv[2])
