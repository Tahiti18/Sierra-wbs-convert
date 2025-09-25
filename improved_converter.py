#!/usr/bin/env python3
"""
Improved Sierra Payroll to WBS Payroll Converter
Handles the exact Sierra format and produces accurate WBS output
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


class SierraToWBSConverter:
    """
    Converts Sierra payroll Excel files to WBS payroll format with accurate calculations
    """
    
    def __init__(self, gold_master_order_path: Optional[str] = None):
        self.gold_master_order = []
        if gold_master_order_path and Path(gold_master_order_path).exists():
            with open(gold_master_order_path, 'r', encoding='utf-8') as f:
                self.gold_master_order = [line.strip() for line in f if line.strip()]
    
    def normalize_name(self, name: str) -> str:
        """Normalize employee name format"""
        if not isinstance(name, str) or not name.strip():
            return ""
        
        # Clean up the name
        name = re.sub(r'\s+', ' ', name.strip())
        
        # If already in "Last, First" format, return as is
        if ',' in name:
            return name
        
        # Convert "First Last" to "Last, First"
        parts = name.split()
        if len(parts) >= 2:
            return f"{parts[-1]}, {' '.join(parts[:-1])}"
        
        return name
    
    def parse_sierra_file(self, file_path: str) -> pd.DataFrame:
        """
        Parse Sierra payroll Excel file and extract employee time data
        """
        # Read the Excel file
        df = pd.read_excel(file_path, header=0)
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Filter out non-data rows
        # Keep rows where Name is not null and Hours > 0
        data_rows = df[
            (df['Name'].notna()) & 
            (df['Name'].astype(str).str.strip() != '') &
            (df['Hours'].notna()) & 
            (pd.to_numeric(df['Hours'], errors='coerce') > 0)
        ].copy()
        
        # Skip signature and other non-employee rows
        skip_keywords = ['signature', 'certify', 'gross', 'week of', 'by the signature']
        for keyword in skip_keywords:
            data_rows = data_rows[
                ~data_rows['Name'].astype(str).str.lower().str.contains(keyword, na=False)
            ]
        
        # Normalize data types
        data_rows['Name'] = data_rows['Name'].astype(str).apply(self.normalize_name)
        data_rows['Hours'] = pd.to_numeric(data_rows['Hours'], errors='coerce').fillna(0.0)
        data_rows['Rate'] = pd.to_numeric(data_rows['Rate'], errors='coerce').fillna(0.0)
        data_rows['Total'] = pd.to_numeric(data_rows['Total'], errors='coerce').fillna(0.0)
        
        # Parse dates
        data_rows['Date'] = pd.to_datetime(data_rows['Days'], errors='coerce')
        
        return data_rows[data_rows['Hours'] > 0].reset_index(drop=True)
    
    def apply_california_overtime_rules(self, employee_data: pd.DataFrame) -> pd.DataFrame:
        """
        Apply California overtime rules:
        - First 8 hours per day: Regular time
        - Hours 8-12 per day: Overtime (1.5x)
        - Hours >12 per day: Double time (2x)
        """
        results = []
        
        # Group by employee and date
        for (name, date_val), group in employee_data.groupby(['Name', 'Date']):
            if pd.isna(date_val):
                # If no date, treat as single day
                daily_hours = group['Hours'].sum()
                rate = group['Rate'].iloc[0]  # Use first rate found
            else:
                daily_hours = group['Hours'].sum()
                rate = group['Rate'].iloc[0]
            
            # Apply CA daily overtime rules
            reg_hours = min(daily_hours, 8.0)
            ot_hours = max(0.0, min(daily_hours - 8.0, 4.0))
            dt_hours = max(0.0, daily_hours - 12.0)
            
            results.append({
                'Name': name,
                'Date': date_val,
                'Rate': rate,
                'REG_Hours': reg_hours,
                'OT_Hours': ot_hours,
                'DT_Hours': dt_hours,
                'Total_Hours': daily_hours
            })
        
        return pd.DataFrame(results)
    
    def aggregate_weekly_data(self, daily_data: pd.DataFrame) -> pd.DataFrame:
        """
        Aggregate daily data to weekly totals per employee
        """
        # Group by employee and sum hours
        weekly = daily_data.groupby('Name').agg({
            'Rate': 'first',  # Use first rate (assuming consistent per employee)
            'REG_Hours': 'sum',
            'OT_Hours': 'sum', 
            'DT_Hours': 'sum',
            'Total_Hours': 'sum'
        }).reset_index()
        
        # Calculate dollar amounts
        weekly['REG_Amount'] = weekly['REG_Hours'] * weekly['Rate']
        weekly['OT_Amount'] = weekly['OT_Hours'] * weekly['Rate'] * 1.5
        weekly['DT_Amount'] = weekly['DT_Hours'] * weekly['Rate'] * 2.0
        weekly['Total_Amount'] = weekly['REG_Amount'] + weekly['OT_Amount'] + weekly['DT_Amount']
        
        return weekly
    
    def sort_employees_by_master_order(self, weekly_data: pd.DataFrame) -> pd.DataFrame:
        """
        Sort employees according to gold master order, with new employees at the end
        """
        if not self.gold_master_order:
            return weekly_data.sort_values('Name')
        
        # Create order mapping
        order_map = {name: i for i, name in enumerate(self.gold_master_order)}
        
        # Add order column
        weekly_data['Order'] = weekly_data['Name'].map(
            lambda x: order_map.get(x, len(self.gold_master_order) + 1000)
        )
        
        # Sort by order, then by name
        return weekly_data.sort_values(['Order', 'Name']).drop('Order', axis=1)
    
    def create_wbs_format(self, weekly_data: pd.DataFrame) -> pd.DataFrame:
        """
        Create WBS format DataFrame with all required columns
        """
        wbs_data = pd.DataFrame()
        
        # Map to WBS column names and add required fields
        wbs_data['SSN'] = ''  # Will be filled from employee database if available
        wbs_data['Employee Name'] = weekly_data['Name']
        wbs_data['Status'] = 'A'  # Active
        wbs_data['Type'] = 'H'    # Hourly (could be 'S' for salary)
        wbs_data['Pay Rate'] = weekly_data['Rate'].round(2)
        wbs_data['Dept'] = ''     # Department - will be filled if available
        wbs_data['A01'] = weekly_data['REG_Hours'].round(2)  # Regular hours
        wbs_data['A02'] = weekly_data['OT_Hours'].round(2)   # Overtime hours
        wbs_data['A03'] = weekly_data['DT_Hours'].round(2)   # Double time hours
        wbs_data['A06'] = 0.0     # Vacation
        wbs_data['A07'] = 0.0     # Sick
        wbs_data['A08'] = 0.0     # Holiday
        wbs_data['A04'] = 0.0     # Bonus
        wbs_data['A05'] = 0.0     # Commission
        wbs_data['REG_$'] = weekly_data['REG_Amount'].round(2)
        wbs_data['OT_$'] = weekly_data['OT_Amount'].round(2)
        wbs_data['DT_$'] = weekly_data['DT_Amount'].round(2)
        wbs_data['TOTAL_$'] = weekly_data['Total_Amount'].round(2)
        
        return wbs_data
    
    def create_wbs_excel(self, wbs_data: pd.DataFrame, output_path: str) -> None:
        """
        Create WBS Excel file with proper formatting and headers
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "WEEKLY"
        
        # WBS Header information
        current_date = datetime.now()
        period_end = current_date.strftime("%m/%d/%Y")
        report_due = current_date.strftime("%m/%d/%Y") 
        check_date = current_date.strftime("%m/%d/%Y")
        
        # Add WBS metadata headers
        headers = [
            ["# V", "DO NOT EDIT", "Version = B90216-00", "FmtRev = 2.1", 
             f"RunTime = {current_date.strftime('%Y%m%d-%H%M%S')}", "CliUnqId = 055269",
             "CliName = Sierra Roofing and Solar Inc", "Freq = W", f"PEDate = {period_end}",
             f"RptDate = {report_due}", f"CkDate = {check_date}", "EmpType = SSN",
             "DoNotes = 1", "PayRates = H+;S+;E+;C+", "RateCol = 6", "T1 = 7+",
             "CodeBeg = 8", "CodeEnd = 26", "NoteCol = 27"],
            ["# U", "CliUnqID", "055269"],
            ["# N", "Client", "Sierra Roofing and Solar Inc"],
            ["# P", "Period End", period_end],
            ["# R", "Report Due", report_due],
            ["# C", "Check Date", check_date],
            ["# T", "EmployeeID", "SSN"],
            ["# B:8", "", "", "", "Pay", "", "", "REGULAR", "OVERTIME", "DOUBLETIME",
             "VACATION", "SICK", "HOLIDAY", "BONUS", "COMMISSION", "PC HRS MON",
             "PC TTL MON", "PC HRS TUE", "PC TTL TUE", "PC HRS WED", "PC TTL WED",
             "PC HRS THU", "PC TTL THU", "PC HRS FRI", "PC TTL FRI", "TRAVEL AMOUNT",
             "Notes and", "Totals"],
            ["# E:26", "SSN", "Employee Name", "Status", "Type", "Pay Rate", "Dept",
             "A01", "A02", "A03", "A06", "A07", "A08", "A04", "A05", "AH1", "AI1",
             "AH2", "AI2", "AH3", "AI3", "AH4", "AI4", "AH5", "AI5", "ATE", "Comments", "Totals"]
        ]
        
        # Write headers
        for header_row in headers:
            ws.append(header_row)
        
        # Write employee data - FIXED COLUMN POSITIONING
        for _, row in wbs_data.iterrows():
            # Calculate total amount for this employee
            total_amount = float(row.get('TOTAL_$', 0) or 0)
            
            # FIXED: Ensure exactly 28 columns with totals in column 27 (index 27)
            employee_row = [
                "",                                          # 0: Employee ID
                row.get('SSN', ''),                         # 1: SSN  
                row.get('Employee Name', ''),               # 2: Employee Name
                row.get('Status', 'A'),                     # 3: Status
                row.get('Type', 'H'),                       # 4: Type
                float(row.get('Pay Rate', 0) or 0),         # 5: Pay Rate
                row.get('Dept', ''),                        # 6: Dept
                float(row.get('A01', 0) or 0),              # 7: A01 - Regular hours
                float(row.get('A02', 0) or 0),              # 8: A02 - OT hours  
                float(row.get('A03', 0) or 0),              # 9: A03 - DT hours
                float(row.get('A06', 0) or 0),              # 10: A06 - Vacation
                float(row.get('A07', 0) or 0),              # 11: A07 - Sick
                float(row.get('A08', 0) or 0),              # 12: A08 - Holiday
                float(row.get('A04', 0) or 0),              # 13: A04 - Bonus
                float(row.get('A05', 0) or 0),              # 14: A05 - Commission
                0,                                          # 15: AH1 - PC HRS MON
                0,                                          # 16: AI1 - PC TTL MON
                0,                                          # 17: AH2 - PC HRS TUE
                0,                                          # 18: AI2 - PC TTL TUE
                0,                                          # 19: AH3 - PC HRS WED
                0,                                          # 20: AI3 - PC TTL WED
                0,                                          # 21: AH4 - PC HRS THU
                0,                                          # 22: AI4 - PC TTL THU
                0,                                          # 23: AH5 - PC HRS FRI
                0,                                          # 24: AI5 - PC TTL FRI
                0,                                          # 25: ATE - TRAVEL AMOUNT
                "",                                         # 26: Comments
                total_amount                                # 27: TOTALS - THIS IS THE FIX!
            ]
            ws.append(employee_row)
        
        # Add totals row - FIXED TO MATCH COLUMN COUNT
        totals_row = [
            "",                                             # 0: Employee ID
            "",                                             # 1: SSN
            "TOTAL",                                        # 2: Employee Name  
            "",                                             # 3: Status
            "",                                             # 4: Type
            "",                                             # 5: Pay Rate
            "",                                             # 6: Dept
            wbs_data['A01'].sum(),                         # 7: Total regular hours
            wbs_data['A02'].sum(),                         # 8: Total OT hours
            wbs_data['A03'].sum(),                         # 9: Total DT hours
            wbs_data['A06'].sum(),                         # 10: Total vacation
            wbs_data['A07'].sum(),                         # 11: Total sick
            wbs_data['A08'].sum(),                         # 12: Total holiday
            wbs_data['A04'].sum(),                         # 13: Total bonus
            wbs_data['A05'].sum(),                         # 14: Total commission
            0,                                             # 15-24: Piecework totals
            0, 0, 0, 0, 0, 0, 0, 0, 0,
            0,                                             # 25: Travel total
            "",                                            # 26: Comments
            wbs_data['TOTAL_$'].sum()                      # 27: GRAND TOTAL
        ]
        ws.append(totals_row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
    
    def convert(self, sierra_file_path: str, output_path: str) -> Dict:
        """
        Main conversion method
        """
        try:
            # Parse Sierra file
            sierra_data = self.parse_sierra_file(sierra_file_path)
            
            if sierra_data.empty:
                return {
                    'success': False,
                    'error': 'No valid employee data found in Sierra file',
                    'employees_processed': 0
                }
            
            # Apply overtime rules
            daily_data = self.apply_california_overtime_rules(sierra_data)
            
            # Aggregate to weekly
            weekly_data = self.aggregate_weekly_data(daily_data)
            
            # Sort by master order
            weekly_data = self.sort_employees_by_master_order(weekly_data)
            
            # Create WBS format
            wbs_data = self.create_wbs_format(weekly_data)
            
            # Create Excel file
            self.create_wbs_excel(wbs_data, output_path)
            
            return {
                'success': True,
                'employees_processed': len(wbs_data),
                'total_hours': weekly_data['Total_Hours'].sum(),
                'total_amount': weekly_data['Total_Amount'].sum(),
                'regular_hours': weekly_data['REG_Hours'].sum(),
                'overtime_hours': weekly_data['OT_Hours'].sum(),
                'doubletime_hours': weekly_data['DT_Hours'].sum()
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'employees_processed': 0
            }


def main():
    """Test the converter"""
    converter = SierraToWBSConverter('/home/ubuntu/upload/sierra_payroll_backend-main/app/data/gold_master_order.txt')
    
    result = converter.convert(
        '/home/ubuntu/upload/SierraPayroll9_19_25forMarwan.xlsx',
        '/home/ubuntu/test_wbs_output.xlsx'
    )
    
    print("Conversion Result:")
    print(f"Success: {result['success']}")
    if result['success']:
        print(f"Employees processed: {result['employees_processed']}")
        print(f"Total hours: {result['total_hours']:.2f}")
        print(f"Regular hours: {result['regular_hours']:.2f}")
        print(f"Overtime hours: {result['overtime_hours']:.2f}")
        print(f"Double time hours: {result['doubletime_hours']:.2f}")
        print(f"Total amount: ${result['total_amount']:.2f}")
    else:
        print(f"Error: {result['error']}")


if __name__ == "__main__":
    main()
