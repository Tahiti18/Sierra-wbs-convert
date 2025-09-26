#!/usr/bin/env python3
"""Fix our WBS converter to match the actual WBS format EXACTLY"""

import openpyxl

print('=== ANALYZING EXACT HEADER STRUCTURE NEEDED ===')

# Load your actual WBS file to get exact format
wb = openpyxl.load_workbook('your_actual_wbs.xlsx')
ws = wb.active

print('CORRECT HEADER STRUCTURE FROM YOUR ACTUAL WBS:')

# Rows 1-8 are all header rows
for row in range(1, 9):
    print(f'\nRow {row}:')
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None and val != "":
            print(f'  Col {col}: "{val}"')

print('\n=== CRITICAL FORMAT REQUIREMENTS ===')
print()

print('🔧 ISSUES TO FIX IN OUR CONVERTER:')
print('1. ❌ WRONG: A01-A03 should contain HOURS, not dollars')
print('2. ❌ WRONG: Totals should be Excel FORMULAS, not calculated values') 
print('3. ❌ WRONG: Empty cells should be None/empty, not zeros')
print('4. ❌ WRONG: Missing Row 7 descriptive headers (Pay, REGULAR, OVERTIME)')
print('5. ❌ WRONG: Department codes (should be ADMIN, GUTTR, not PRODUCTION)')
print('6. ❌ WRONG: Missing proper two-line header structure')

print()
print('✅ CORRECT FORMAT REQUIREMENTS:')
print('- A01: Regular HOURS (like 4.0, 32.0, 40.0)')
print('- A02: Overtime HOURS (like 6.0, 0.5)')  
print('- A03: Double-time HOURS (rare, usually None)')
print('- Total: Excel formula like =(F9*H9)+(F9*I9)+...')
print('- Empty cells: None/blank, NOT zeros')
print('- Departments: ADMIN, GUTTR, etc.')
print('- Two header rows: Row 7 (descriptive) + Row 8 (codes)')

print()
print('🚨 CRITICAL: The payroll company expects the ORIGINAL format!')
print('   They want hours + formulas, NOT calculated dollar amounts!')

# Generate the exact header structure needed
print('\n=== EXACT HEADER STRUCTURE NEEDED ===')

headers_needed = []
for row in range(1, 9):
    row_data = []
    for col in range(1, 29):  # 28 columns total
        val = ws.cell(row=row, column=col).value
        row_data.append(val)
    headers_needed.append(row_data)
    
print('Copy this exact header structure:')
for i, header_row in enumerate(headers_needed, 1):
    print(f'Row {i}: {header_row}')