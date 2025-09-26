#!/usr/bin/env python3
"""
Extract complete employee database from actual WBS file
"""

import openpyxl

def extract_complete_employee_database():
    """Extract all employees from actual WBS file"""
    wb = openpyxl.load_workbook('your_actual_wbs.xlsx')
    ws = wb.active
    
    employees = {}
    for row in range(9, ws.max_row + 1):
        emp_num = ws.cell(row, 1).value
        ssn = ws.cell(row, 2).value 
        name = ws.cell(row, 3).value
        status = ws.cell(row, 4).value
        emp_type = ws.cell(row, 5).value
        dept = ws.cell(row, 7).value
        
        if name and emp_num and ssn:
            employees[name] = {
                'employee_number': str(emp_num).zfill(10),
                'ssn': str(ssn),
                'status': status or 'A',
                'type': emp_type or 'H', 
                'department': dept or 'UNKNOWN'
            }
    
    wb.close()
    
    # Generate Python code for the database
    print('    def _create_employee_database(self) -> Dict[str, Dict]:')
    print('        """Create employee database with exact WBS format data"""')
    print('        return {')
    
    for name, info in sorted(employees.items()):
        print(f'            "{name}": {{')
        print(f'                "employee_number": "{info["employee_number"]}",')
        print(f'                "ssn": "{info["ssn"]}",')
        print(f'                "status": "{info["status"]}",')
        print(f'                "type": "{info["type"]}",')
        print(f'                "department": "{info["department"]}"')
        print('            },')
    
    print('        }')
    
    return employees

if __name__ == '__main__':
    employees = extract_complete_employee_database()
    print(f'\n# Total employees extracted: {len(employees)}')