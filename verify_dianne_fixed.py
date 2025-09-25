#!/usr/bin/env python3
"""
Quick verification of Dianne's data in fixed output
"""
import pandas as pd
from openpyxl import load_workbook

def verify_dianne():
    print("=== VERIFYING DIANNE'S FIXED OUTPUT ===")
    
    # Load fixed output 
    wb = load_workbook('/home/user/webapp/wbs_output_FIXED.xlsx')
    ws = wb.active
    
    # Dianne is in row 24 according to our search
    dianne_row = 24
    
    print(f"Dianne's data (Row {dianne_row}):")
    
    # Extract all column values
    dianne_data = []
    for col in range(1, 29):
        value = ws.cell(row=dianne_row, column=col).value
        dianne_data.append(value)
    
    # Key fields
    print(f"  Employee Number: {dianne_data[0]}")
    print(f"  SSN: {dianne_data[1]}")
    print(f"  Name: {dianne_data[2]}")
    print(f"  Status: {dianne_data[3]}")
    print(f"  Type: {dianne_data[4]}")
    print(f"  Rate: ${dianne_data[5]}")
    print(f"  Department: {dianne_data[6]}")
    print(f"  Regular Hours (A01): {dianne_data[7]}")
    print(f"  OT1.5 Hours (A02): {dianne_data[8]}")
    print(f"  OT2.0 Hours (A03): {dianne_data[9]}")
    print(f"  Total Amount: ${dianne_data[27]}")
    
    # Verification
    print(f"\n=== VERIFICATION AGAINST GOLD STANDARD ===")
    expected_emp_num = "0000662082"
    expected_ssn = "626946016"
    expected_rate = 28.0
    expected_hours = 4.0
    expected_total = 112.0
    
    # Convert for comparison
    actual_emp_num = str(int(dianne_data[0])).zfill(10) if dianne_data[0] else "0"
    actual_ssn = str(int(dianne_data[1])) if dianne_data[1] else "0"
    actual_rate = float(dianne_data[5]) if dianne_data[5] else 0
    actual_hours = float(dianne_data[7]) if dianne_data[7] else 0
    actual_total = float(dianne_data[27]) if dianne_data[27] else 0
    
    emp_match = actual_emp_num == expected_emp_num
    ssn_match = actual_ssn == expected_ssn
    rate_match = abs(actual_rate - expected_rate) < 0.01
    hours_match = abs(actual_hours - expected_hours) < 0.01
    total_match = abs(actual_total - expected_total) < 0.01
    name_match = "Dianne" in str(dianne_data[2])
    
    print(f"Employee Number: {'✅' if emp_match else '❌'} (Expected: {expected_emp_num}, Got: {actual_emp_num})")
    print(f"SSN: {'✅' if ssn_match else '❌'} (Expected: {expected_ssn}, Got: {actual_ssn})")
    print(f"Rate: {'✅' if rate_match else '❌'} (Expected: ${expected_rate}, Got: ${actual_rate})")
    print(f"Hours: {'✅' if hours_match else '❌'} (Expected: {expected_hours}, Got: {actual_hours})")
    print(f"Total: {'✅' if total_match else '❌'} (Expected: ${expected_total}, Got: ${actual_total})")
    print(f"Name: {'✅' if name_match else '❌'} (Contains 'Dianne': {name_match})")
    
    all_match = emp_match and ssn_match and rate_match and hours_match and total_match and name_match
    
    print(f"\n🎯 OVERALL: {'🎉 PERFECT! All data matches gold standard!' if all_match else '❌ Issues found'}")
    
    # Check if values are calculated (not formulas)
    is_calculated = not (isinstance(dianne_data[27], str) and dianne_data[27].startswith('='))
    print(f"Total is calculated value (not formula): {'✅' if is_calculated else '❌'}")
    
    # Check for None values that should be 0
    none_count = sum(1 for val in dianne_data[10:26] if val is None)  # Middle columns should be 0
    print(f"Middle columns properly filled (not None): {'✅' if none_count == 0 else f'❌ {none_count} None values'}")

if __name__ == "__main__":
    verify_dianne()