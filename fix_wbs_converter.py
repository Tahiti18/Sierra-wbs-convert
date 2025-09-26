#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook

def analyze_actual_wbs_structure():
    """Analyze how the actual WBS file stores amounts"""
    
    print("ANALYZING ACTUAL WBS STRUCTURE")
    print("=" * 80)
    
    # Load the actual WBS file with openpyxl to see formulas
    wb = load_workbook("/home/user/webapp/WBS_Payroll_9_12_25_for_Marwan_2.xlsx")
    ws = wb.active
    
    print("Checking Dianne's row (row 8)...")
    
    # Check Dianne's row structure (row 8 based on our analysis)
    dianne_row = 8
    
    for col in range(1, 29):  # Check all columns
        cell = ws.cell(row=dianne_row, column=col)
        if cell.value is not None:
            print(f"Col {col}: '{cell.value}' (Type: {type(cell.value)})")
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                print(f"  Formula: {cell.value}")
    
    print("\nChecking our generated WBS structure...")
    
    # Load our generated file
    our_wb = load_workbook("/home/user/webapp/test_new_conversion.xlsx")
    our_ws = our_wb.active
    
    # Find Dianne in our file (should be row 9 based on headers)
    for row in range(1, 20):
        cell = our_ws.cell(row=row, column=3)  # Name column
        if cell.value and 'Dianne' in str(cell.value):
            print(f"\nFound Dianne in our file at row {row}")
            
            for col in range(1, 29):
                our_cell = our_ws.cell(row=row, column=col)
                if our_cell.value is not None:
                    print(f"Col {col}: '{our_cell.value}' (Type: {type(our_cell.value)})")
            break

def check_excel_formula_issue():
    """Test if the Excel formulas are working correctly"""
    
    print("\n" + "=" * 80)
    print("CHECKING EXCEL FORMULA CALCULATION")
    print("=" * 80)
    
    # Read our generated file as pandas to see calculated values
    df = pd.read_excel("/home/user/webapp/test_new_conversion.xlsx", sheet_name="WEEKLY")
    
    print("Our generated file structure:")
    print(f"Shape: {df.shape}")
    
    # Find employee rows (those with SSNs)
    for idx, row in df.iterrows():
        name = str(row.iloc[2]) if len(row) > 2 else ""
        if 'Dianne' in name:
            print(f"\nDianne row {idx}:")
            for i, val in enumerate(row):
                if pd.notna(val):
                    print(f"  Col {i}: {val}")
            break

if __name__ == "__main__":
    analyze_actual_wbs_structure()
    check_excel_formula_issue()