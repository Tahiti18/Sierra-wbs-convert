#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook

def analyze_gold_standard():
    """Analyze the gold standard WBS file to get exact requirements"""
    
    print("GOLD STANDARD WBS ANALYSIS")
    print("=" * 80)
    
    # Load the gold standard with calculated values
    wb = load_workbook('/home/user/webapp/WBS_Payroll_9_12_25_for_Marwan.xlsx', data_only=True)
    ws = wb.active
    
    print("Extracting all employees in exact order:")
    print("-" * 50)
    
    employees = []
    total_amount = 0
    
    for row_num in range(9, 200):  # Check many rows to get all employees
        name_cell = ws.cell(row=row_num, column=3)
        ssn_cell = ws.cell(row=row_num, column=2)
        total_cell = ws.cell(row=row_num, column=28)
        
        if name_cell.value and isinstance(name_cell.value, str) and len(name_cell.value) > 3:
            # This is an employee row
            employee_name = name_cell.value
            ssn = ssn_cell.value if ssn_cell.value else "No SSN"
            amount = total_cell.value if total_cell.value else 0
            
            employees.append({
                'position': len(employees) + 1,
                'name': employee_name,
                'ssn': ssn,
                'amount': amount
            })
            
            total_amount += amount if amount else 0
            
            print(f"{len(employees):3d}. {employee_name:30} | SSN: {ssn:12} | ${amount}")
    
    print("-" * 80)
    print(f"TOTAL EMPLOYEES: {len(employees)}")
    print(f"TOTAL AMOUNT: ${total_amount:,.2f}")
    
    # Save the exact order for our converter
    with open('/home/user/webapp/gold_standard_order.txt', 'w') as f:
        f.write("# EXACT WBS Employee Order - Gold Standard\n")
        f.write(f"# Total Employees: {len(employees)}\n")
        f.write(f"# Total Amount: ${total_amount:,.2f}\n\n")
        
        for emp in employees:
            f.write(f'"{emp["name"]}",  # {emp["position"]:3d} - ${emp["amount"]}\n')
    
    print(f"\nGold standard order saved to: gold_standard_order.txt")
    
    # Find employees with amounts > 0 (worked this week)
    working_employees = [emp for emp in employees if emp['amount'] > 0]
    print(f"\nEmployees who worked this week: {len(working_employees)}")
    
    return employees

if __name__ == "__main__":
    employees = analyze_gold_standard()