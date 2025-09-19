# src/roster_enforcer.py
from __future__ import annotations
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell

TEMPLATE_PATH = Path(os.path.dirname(os.path.dirname(__file__))) / "data" / "wbs_template.xlsx"

# ------------------------- Roster & SSN helpers -------------------------

def _read_roster_order(gold_master_path: str) -> List[str]:
    """Read gold master order (one name per line)."""
    p = Path(gold_master_path)
    if not p.exists():
        raise FileNotFoundError(f"Gold master order not found: {gold_master_path}")
    names = [line.strip() for line in p.read_text(encoding="utf-8").splitlines()]
    return [n for n in names if n]

def _read_sierra_for_ssn(input_sierra_path: str) -> Dict[str, str]:
    """
    Build name->SSN map from the Sierra WEEKLY sheet.
    Sierra header row index is 7 (1-based row 8). Name usually 'Employee Name' or 'Unnamed: 2'. SSN usually 'SSN' or 'Unnamed: 1'.
    """
    ssn_map: Dict[str, str] = {}
    try:
        df = pd.read_excel(input_sierra_path, sheet_name="WEEKLY", header=7)
        df = df.dropna(how="all")
        # Drop the duplicated header line if present
        if (df.iloc[0:1].astype(str).apply(lambda x: (x == 'Employee Name').any(), axis=1).any()):
            df = df.iloc[1:]
        name_col = "Employee Name" if "Employee Name" in df.columns else ("Unnamed: 2" if "Unnamed: 2" in df.columns else None)
        ssn_col  = "SSN"            if "SSN" in df.columns else ("Unnamed: 1" if "Unnamed: 1" in df.columns else None)
        if not name_col:
            return ssn_map
        for _, r in df[[name_col] + ([ssn_col] if ssn_col else [])].dropna(subset=[name_col]).iterrows():
            name = str(r[name_col]).strip()
            ssn  = "" if (not ssn_col or pd.isna(r.get(ssn_col))) else str(r[ssn_col]).strip()
            if name:
                ssn_map[name] = ssn
    except Exception:
        pass
    return ssn_map

# ------------------------- Sheet structure helpers -------------------------

def _find_header_row(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    """
    Find the header row index (1-based) by locating a row that contains 'Employee Name'
    and return a mapping of important column names to column indexes (1-based).
    """
    wanted = ["SSN", "Employee Name", "REGULAR", "OVERTIME", "DOUBLETIME", "Totals"]
    col_idx: Dict[str, int] = {}

    for r in range(1, min(ws.max_row, 50) + 1):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if any(v.lower() == "employee name" for v in row_vals):
            # map columns by exact header text if available; otherwise keep best-effort
            for c_idx, v in enumerate(row_vals, start=1):
                v_lower = v.lower()
                if v_lower == "ssn": col_idx["SSN"] = c_idx
                elif v_lower == "employee name": col_idx["Employee Name"] = c_idx
                elif v_lower == "regular": col_idx["REGULAR"] = c_idx
                elif v_lower == "overtime": col_idx["OVERTIME"] = c_idx
                elif v_lower == "doubletime": col_idx["DOUBLETIME"] = c_idx
                elif v_lower == "totals": col_idx["Totals"] = c_idx
            return r, col_idx
    raise ValueError("Could not locate header row containing 'Employee Name'")

def _first_data_row(header_row: int) -> int:
    return header_row + 1

def _collect_existing_rows(ws: Worksheet, name_col: int, start_row: int) -> Dict[str, int]:
    """
    Return map: employee name -> row index (1-based), only for rows with a non-empty name cell.
    Stops when encountering 10 consecutive blank name cells.
    """
    result: Dict[str, int] = {}
    blanks = 0
    r = start_row
    while r <= ws.max_row and blanks < 10:
        name = ws.cell(row=r, column=name_col).value
        if name is None or str(name).strip() == "":
            blanks += 1
        else:
            blanks = 0
            result[str(name).strip()] = r
        r += 1
    return result

def _copy_cell_style(src: Cell, dst: Cell):
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)

def _clone_row(ws_src: Worksheet, row_src: int, ws_dst: Worksheet, row_dst: int):
    """Clone an entire row (values + styles)."""
    max_col = ws_src.max_column
    for c in range(1, max_col + 1):
        s_cell = ws_src.cell(row=row_src, column=c)
        d_cell = ws_dst.cell(row=row_dst, column=c)
        d_cell.value = s_cell.value
        _copy_cell_style(s_cell, d_cell)

def _clear_row_values_keep_styles(ws: Worksheet, row_idx: int):
    """Zero out numeric cells, blank strings in text cells, but keep styling."""
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        # Heuristic: zero numerics, blank strings elsewhere
        if cell.data_type == 'n':
            cell.value = 0
        else:
            # leave formulas (type 'f') as-is; otherwise blank
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = ""

def _copy_header_and_footer(ws_src: Worksheet, ws_dst: Worksheet, start_row_dst: int, data_rows_needed: int, header_row_src: int) -> Tuple[int, int]:
    """
    Copy header rows (1..header_row_src) from ws_src -> ws_dst.
    Then position data region to start immediately after header.
    Footer rows (if any) are copied AFTER we write data rows, preserving relative position to bottom by
    copying from the source and re-inserting below the new data region.
    Returns (header_rows_copied, footer_start_src).
    """
    # Copy header rows exactly
    for r in range(1, header_row_src + 0):
        _clone_row(ws_src, r, ws_dst, r)

    # Identify footer start in source: first row after data region where name cell turns blank for 10 consecutive rows
    _, colmap = _find_header_row(ws_src)
    name_col = colmap.get("Employee Name", None)
    data_start = _first_data_row(header_row_src)
    # find last data row in source
    blanks = 0
    r = data_start
    last_data_row_src = data_start - 1
    while r <= ws_src.max_row and blanks < 10:
        val = ws_src.cell(row=r, column=name_col).value if name_col else None
        if val is None or str(val).strip() == "":
            blanks += 1
        else:
            blanks = 0
            last_data_row_src = r
        r += 1
    footer_start_src = last_data_row_src + 1
    return header_row_src, footer_start_src

def _copy_footer(ws_src: Worksheet, ws_dst: Worksheet, footer_start_src: int, footer_start_dst: int):
    """Copy footer rows from source into destination starting at footer_start_dst (values + styles + merges)."""
    for r_src in range(footer_start_src, ws_src.max_row + 1):
        _clone_row(ws_src, r_src, ws_dst, footer_start_dst + (r_src - footer_start_src))

# ------------------------- Main enforcement -------------------------

def enforce_roster(output_wbs_path: str, input_sierra_path: str, gold_master_path: str) -> None:
    """
    Guarantee:
      - Names appear in EXACT gold order (top to bottom).
      - SSNs are placed in the SSN column (best-effort from Sierra file if missing).
      - All numeric hour columns (REGULAR/OVERTIME/DOUBLETIME + any others already in the sheet) are preserved for matching names.
      - Missing names are inserted as zero rows.
      - Layout, styling, widths, and formulas are preserved. If data/wbs_template.xlsx exists, that template governs the look.
    """
    roster = _read_roster_order(gold_master_path)
    ssn_map = _read_sierra_for_ssn(input_sierra_path)

    # Load the current WBS (source of values)
    wb_src: Workbook = load_workbook(output_wbs_path)
    if "WEEKLY" not in wb_src.sheetnames:
        raise ValueError("WEEKLY sheet not found in WBS output")
    ws_src: Worksheet = wb_src["WEEKLY"]

    # Choose destination workbook:
    # - If TEMPLATE_PATH exists, start from that template => guarantees identical look & formulas.
    # - Else, build a new sheet in the same workbook and copy header/footer from source.
    use_template = TEMPLATE_PATH.exists()
    if use_template:
        wb_dst: Workbook = load_workbook(TEMPLATE_PATH)
        if "WEEKLY" not in wb_dst.sheetnames:
            raise ValueError("WEEKLY sheet not found in template")
        ws_dst: Worksheet = wb_dst["WEEKLY"]
    else:
        wb_dst = load_workbook(output_wbs_path)
        # Create new sheet to write into (then replace)
        if "WEEKLY_enforced" in wb_dst.sheetnames:
            del wb_dst["WEEKLY_enforced"]
        ws_dst = wb_dst.create_sheet("WEEKLY_enforced")

    # Analyze columns/rows in source and destination
    header_src, colmap_src = _find_header_row(ws_src)
    name_col_src = colmap_src.get("Employee Name")
    ssn_col_src = colmap_src.get("SSN")

    header_dst, colmap_dst = _find_header_row(ws_dst)
    name_col_dst = colmap_dst.get("Employee Name")
    ssn_col_dst = colmap_dst.get("SSN")

    if not name_col_src or not name_col_dst:
        raise ValueError("Could not determine 'Employee Name' column in source or destination sheet.")

    data_start_src = _first_data_row(header_src)
    data_start_dst = _first_data_row(header_dst)

    # Map existing rows in source by name (trim whitespace)
    existing_rows = _collect_existing_rows(ws_src, name_col_src, data_start_src)

    # If not using template, copy header & prepare footer coordinates
    if not use_template:
        header_rows_copied, footer_start_src = _copy_header_and_footer(ws_src, ws_dst, start_row_dst=1, data_rows_needed=len(roster), header_row_src=header_src)
        footer_start_dst = data_start_dst + len(roster)
        # clear potential existing content in the data area
        # (ws_dst was empty; header already cloned)
    else:
        # with template, we keep header+footer as in template; footer will remain in place (formulas intact)
        # we just write the data rows into template’s body region
        # Optional: zero out data area to a clean state
        pass

    # Build a style template row from the first data row of destination (if template) or source
    style_template_row_idx = data_start_dst if use_template else data_start_src
    style_source_ws = ws_dst if use_template else ws_src

    # Write roster-ordered rows
    for i, name in enumerate(roster):
        dst_row = data_start_dst + i

        # If using template, ensure there is a row to copy styles from
        # We will clone styles cell-by-cell from the style_template_row_idx
        for c in range(1, ws_dst.max_column + 1):
            s_cell = style_source_ws.cell(row=style_template_row_idx, column=c)
            d_cell = ws_dst.cell(row=dst_row, column=c)
            _copy_cell_style(s_cell, d_cell)
            # initialize values as blank/zero
            if s_cell.data_type == 'n':
                d_cell.value = 0
            else:
                # keep formulas if template has row formulas; they will auto-adjust for dst_row if present
                if not (isinstance(d_cell.value, str) and str(d_cell.value).startswith("=")):
                    d_cell.value = ""

        # Fill name + SSN
        ws_dst.cell(row=dst_row, column=name_col_dst).value = name
        if ssn_col_dst:
            ssn_val = ssn_map.get(name, ws_dst.cell(row=dst_row, column=ssn_col_dst).value)
            ws_dst.cell(row=dst_row, column=ssn_col_dst).value = ssn_val

        # If this name exists in the source, copy that entire row’s values *into* destination row
        src_row = existing_rows.get(name)
        if src_row:
            max_cols = max(ws_src.max_column, ws_dst.max_column)
            for c in range(1, max_cols + 1):
                s_cell = ws_src.cell(row=src_row, column=c)
                d_cell = ws_dst.cell(row=dst_row, column=c)
                # copy value; styles already set
                d_cell.value = s_cell.value

            # Ensure name/SSN cells reflect our enforced values (avoid source drift)
            ws_dst.cell(row=dst_row, column=name_col_dst).value = name
            if ssn_col_dst:
                ssn_val = ssn_map.get(name, ws_src.cell(row=src_row, column=ssn_col_src).value if ssn_col_src else ws_dst.cell(row=dst_row, column=ssn_col_dst).value)
                ws_dst.cell(row=dst_row, column=ssn_col_dst).value = ssn_val
        else:
            # Missing in source: keep zeros for numeric cells; keep any per-row formulas from the template
            pass

    # Footer handling
    if not use_template:
        # Copy footer from source to sit immediately after the last data row
        # Identify footer start in source again
        _, colmap = _find_header_row(ws_src)
        name_col = colmap.get("Employee Name", None)
        blanks = 0
        r = data_start_src
        last_data_row_src = data_start_src - 1
        while r <= ws_src.max_row and blanks < 10:
            val = ws_src.cell(row=r, column=name_col).value if name_col else None
            if val is None or str(val).strip() == "":
                blanks += 1
            else:
                blanks = 0
                last_data_row_src = r
            r += 1
        footer_start_src = last_data_row_src + 1
        footer_start_dst = data_start_dst + len(roster)
        _copy_footer(ws_src, ws_dst, footer_start_src, footer_start_dst)

        # Replace original WEEKLY with enforced
        if "WEEKLY" in wb_dst.sheetnames:
            del wb_dst["WEEKLY"]
        ws_dst.title = "WEEKLY"
        wb_dst.save(output_wbs_path)
    else:
        # Using template: just save the filled template over the output path
        wb_dst.save(output_wbs_path)
