#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook

def find_dianne_in_actual_wbs():
    """Find Dianne's actual row in the WBS file"""
    
    print("FINDING DIANNE'S CORRECT ROW")
    print("=" * 80)
    
    # Load with pandas first
    df = pd.read_excel("/home/user/webapp/WBS_Payroll_9_12_25_for_Marwan_2.xlsx", sheet_name="WEEKLY")
    
    print("Looking for Dianne in pandas dataframe...")
    for idx, row in df.iterrows():
        name = str(row.iloc[2]) if len(row) > 2 else ""
        if 'Dianne' in name:
            print(f"Found Dianne at pandas row {idx}:")
            print(f"  Name: {name}")
            print(f"  Hours (col 7): {row.iloc[7] if len(row) > 7 else 'N/A'}")
            print(f"  Total (col 27): {row.iloc[27] if len(row) > 27 else 'N/A'}")
            
            # Now check with openpyxl (row numbers are +1)
            wb = load_workbook("/home/user/webapp/WBS_Payroll_9_12_25_for_Marwan_2.xlsx")
            ws = wb.active
            
            excel_row = idx + 1
            print(f"\nOpenpyxl row {excel_row} (pandas row {idx}):")
            
            # Check specific columns
            name_cell = ws.cell(row=excel_row, column=3)
            hours_cell = ws.cell(row=excel_row, column=8)  # A01 column
            rate_cell = ws.cell(row=excel_row, column=6)
            total_cell = ws.cell(row=excel_row, column=28)  # Totals column
            
            print(f"  Name: '{name_cell.value}'")
            print(f"  Rate: '{rate_cell.value}'") 
            print(f"  Hours: '{hours_cell.value}'")
            print(f"  Total: '{total_cell.value}' (Type: {type(total_cell.value)})")
            
            # Check if total is a formula
            if hasattr(total_cell, 'data_type') and total_cell.data_type == 'f':
                print(f"  Total Formula: {total_cell.value}")
            
            break

def check_formula_structure():
    """Check how the actual WBS calculates totals"""
    
    print("\n" + "=" * 80)
    print("ANALYZING TOTAL CALCULATION STRUCTURE")
    print("=" * 80)
    
    wb = load_workbook("/home/user/webapp/WBS_Payroll_9_12_25_for_Marwan_2.xlsx")
    ws = wb.active
    
    # Check several employee rows to understand the pattern
    dianne_row = 8  # Based on our earlier analysis, Dianne might be at row 8
    
    print("Checking various rows for formula patterns...")
    
    for row_num in range(7, 15):  # Check rows 7-14
        name_cell = ws.cell(row=row_num, column=3)
        total_cell = ws.cell(row=row_num, column=28)
        
        if name_cell.value and isinstance(name_cell.value, str) and len(str(name_cell.value)) > 5:
            print(f"\nRow {row_num}: {name_cell.value}")
            print(f"  Total: {total_cell.value} (Type: {type(total_cell.value)})")
            
            # If it's a number, it's likely calculated already
            if isinstance(total_cell.value, (int, float)):
                print(f"  Calculated value: ${total_cell.value}")
                
                # Check the rate and hours to verify calculation
                rate_cell = ws.cell(row=row_num, column=6)
                hours_cell = ws.cell(row=row_num, column=8)
                
                print(f"  Rate: {rate_cell.value}, Hours: {hours_cell.value}")
                
                if rate_cell.value and hours_cell.value:
                    expected = float(rate_cell.value) * float(hours_cell.value)
                    print(f"  Expected (rate * hours): ${expected}")

if __name__ == "__main__":
    find_dianne_in_actual_wbs()
    check_formula_structure()