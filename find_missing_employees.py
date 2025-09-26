#!/usr/bin/env python3
"""Find employees missing from our database"""

import openpyxl
from wbs_accurate_converter import WBSAccurateConverter

# Load the converter to see current employee database
converter = WBSAccurateConverter()
print('=== CURRENT EMPLOYEE DATABASE ===')
print(f'Database has {len(converter.employee_database)} employees:')
for name, info in converter.employee_database.items():
    print(f'  {name}: {info["employee_number"]} (SSN: {info["ssn"]})')

# Load the WBS conversion result
wb = openpyxl.load_workbook('WBS_Full_Conversion_Test.xlsx')
ws = wb.active

print('\n=== EMPLOYEES IN CONVERSION RESULT ===')
missing_employees = []
found_employees = []

for row_num in range(9, ws.max_row + 1):
    name = ws.cell(row=row_num, column=3).value
    employee_num = ws.cell(row=row_num, column=1).value
    ssn = ws.cell(row=row_num, column=2).value
    
    if name and name.strip():
        if name in converter.employee_database:
            found_employees.append(name)
            print(f'✅ {name}: Found in database')
        else:
            missing_employees.append(name)
            print(f'❌ {name}: MISSING from database (assigned {employee_num}, SSN: {ssn})')

print(f'\n=== SUMMARY ===')
print(f'✅ Found in database: {len(found_employees)} employees')
print(f'❌ Missing from database: {len(missing_employees)} employees')

if missing_employees:
    print(f'\n=== MISSING EMPLOYEES LIST ===')
    for i, name in enumerate(missing_employees, 1):
        print(f'{i:2d}. {name}')
        
    print(f'\n🎯 NEXT STEP: Add these {len(missing_employees)} employees to the employee database')
    print('   with proper employee numbers and SSNs for complete accuracy.')