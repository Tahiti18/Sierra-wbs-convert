#!/usr/bin/env python3
"""Extract correct departments from actual WBS file"""

import openpyxl

wb = openpyxl.load_workbook('your_actual_wbs.xlsx')
ws = wb.active

print('=== EXTRACTING CORRECT DEPARTMENTS ===')

dept_mapping = {}
for row in range(9, min(50, ws.max_row + 1)):  # First 40 employees
    name = ws.cell(row=row, column=3).value
    dept = ws.cell(row=row, column=7).value
    emp_num = ws.cell(row=row, column=1).value
    ssn = ws.cell(row=row, column=2).value
    
    if name and dept and emp_num and ssn:
        # Clean up name format
        name = str(name).strip()
        dept = str(dept).strip()
        
        dept_mapping[name] = {
            'department': dept,
            'employee_number': str(emp_num),
            'ssn': str(ssn)
        }
        
        print(f'{name}: {dept} (#{emp_num}, SSN: {ssn})')

print(f'\nFound {len(dept_mapping)} employees with correct departments')

print('\n=== DEPARTMENT CODES FOUND ===')
departments = set()
for info in dept_mapping.values():
    departments.add(info['department'])
    
for dept in sorted(departments):
    print(f'- {dept}')

print('\n=== UPDATED EMPLOYEE DATABASE NEEDED ===')
print('Update these employees with correct departments:')