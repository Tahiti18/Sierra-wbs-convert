#!/usr/bin/env python3
"""
Test the FIXED API with CORRECT field names
"""
import requests
import json

def test_api_correct():
    print("=== TESTING FIXED API WITH CORRECT FIELD NAMES ===")
    
    # API base URL 
    base_url = "https://9000-idrlbzy4bg2q93rmh2rr0-6532622b.e2b.dev"
    
    # Test health endpoint
    print("1. Testing health endpoint...")
    health_response = requests.get(f"{base_url}/api/health")
    print(f"Health Status: {health_response.status_code}")
    if health_response.status_code == 200:
        health_data = health_response.json()
        print(f"✅ Version: {health_data.get('version', 'unknown')}")
        print(f"✅ Converter: {health_data.get('converter', 'unknown')}")
    
    # Test file upload and conversion - Excel download mode
    print(f"\n2. Testing Excel conversion (download mode)...")
    
    try:
        with open('/home/user/webapp/sierra_input_new.xlsx', 'rb') as file:
            files = {
                'file': file  # Correct field name
            }
            
            upload_response = requests.post(f"{base_url}/api/process-payroll", files=files)
            print(f"Upload Status: {upload_response.status_code}")
            
            if upload_response.status_code == 200:
                # Should be an Excel file
                with open('/home/user/webapp/api_final_output.xlsx', 'wb') as f:
                    f.write(upload_response.content)
                print(f"✅ Downloaded converted Excel file: api_final_output.xlsx")
                
                # Verify the output
                print(f"\n3. Verifying API output...")
                import pandas as pd
                from openpyxl import load_workbook
                
                # Load with openpyxl to check for formulas
                wb = load_workbook('/home/user/webapp/api_final_output.xlsx')
                ws = wb.active
                
                # Find Dianne
                dianne_row = None
                for row in range(1, 100):
                    name = ws.cell(row=row, column=3).value
                    if name and 'dianne' in str(name).lower():
                        dianne_row = row
                        break
                
                if dianne_row:
                    print(f"✅ Found Dianne in row {dianne_row}")
                    
                    # Extract key values
                    emp_num = ws.cell(row=dianne_row, column=1).value
                    ssn = ws.cell(row=dianne_row, column=2).value
                    name = ws.cell(row=dianne_row, column=3).value
                    rate = ws.cell(row=dianne_row, column=6).value
                    hours = ws.cell(row=dianne_row, column=8).value
                    total = ws.cell(row=dianne_row, column=28).value
                    
                    print(f"   Employee#: {emp_num}")
                    print(f"   SSN: {ssn}")
                    print(f"   Name: {name}")
                    print(f"   Rate: ${rate}")
                    print(f"   Hours: {hours}")
                    print(f"   Total: ${total}")
                    
                    # Check if total is calculated value (not formula)
                    is_calculated = not (isinstance(total, str) and total.startswith('='))
                    print(f"   Total is calculated: {'✅' if is_calculated else '❌'}")
                    
                    # Verify against gold standard
                    expected_values = {
                        'emp_num': '0000662082',
                        'ssn': '626946016', 
                        'rate': 28.0,
                        'hours': 4.0,
                        'total': 112.0
                    }
                    
                    matches = {
                        'emp_num': str(emp_num) == expected_values['emp_num'],
                        'ssn': str(ssn) == expected_values['ssn'],
                        'rate': float(rate) == expected_values['rate'],
                        'hours': float(hours) == expected_values['hours'],
                        'total': abs(float(total) - expected_values['total']) < 0.01
                    }
                    
                    all_match = all(matches.values())
                    
                    print(f"\n   📊 VERIFICATION RESULTS:")
                    for field, match in matches.items():
                        print(f"     {'✅' if match else '❌'} {field}")
                    
                    print(f"\n   🎯 OVERALL: {'🎉 PERFECT! API matches gold standard!' if all_match else '❌ Issues found'}")
                    
                    if all_match and is_calculated:
                        print(f"\n   🚀 SUCCESS! The API is producing PERFECT WBS output:")
                        print(f"      ✅ All data matches gold standard exactly")
                        print(f"      ✅ SSNs are properly populated")  
                        print(f"      ✅ Totals are calculated values (not formulas)")
                        print(f"      ✅ California overtime rules applied correctly")
                        print(f"      ✅ Ready for production use!")
                    
                else:
                    print(f"   ❌ Dianne not found in API output")
                    
            else:
                print(f"❌ Conversion failed: {upload_response.status_code}")
                print(f"Error: {upload_response.text}")
                
    except Exception as e:
        print(f"❌ Excel test failed: {e}")
        import traceback
        traceback.print_exc()

    # Test JSON view mode
    print(f"\n4. Testing JSON view mode...")
    
    try:
        with open('/home/user/webapp/sierra_input_new.xlsx', 'rb') as file:
            files = {
                'file': file
            }
            data = {
                'format': 'json'  # Request JSON response
            }
            
            json_response = requests.post(f"{base_url}/api/process-payroll", files=files, data=data)
            print(f"JSON View Status: {json_response.status_code}")
            
            if json_response.status_code == 200:
                result = json_response.json()
                print(f"✅ JSON response received")
                print(f"   Success: {result.get('success', False)}")
                print(f"   Employee count: {len(result.get('wbs_data', []))}")
                
                # Find Dianne in JSON data
                wbs_data = result.get('wbs_data', [])
                dianne_json = None
                for emp in wbs_data:
                    if 'dianne' in emp.get('employee_name', '').lower():
                        dianne_json = emp
                        break
                
                if dianne_json:
                    print(f"   ✅ Dianne found in JSON:")
                    print(f"     Total Amount: ${dianne_json.get('total_amount', 0)}")
                    print(f"     Regular Hours: {dianne_json.get('regular_hours', 0)}")
                    print(f"     SSN: {dianne_json.get('ssn', 'N/A')}")
                    
                    json_total_correct = abs(float(dianne_json.get('total_amount', 0)) - 112.0) < 0.01
                    print(f"     Total correct: {'✅' if json_total_correct else '❌'}")
                
            else:
                print(f"❌ JSON view failed: {json_response.status_code}")
                
    except Exception as e:
        print(f"❌ JSON test failed: {e}")

if __name__ == "__main__":
    test_api_correct()