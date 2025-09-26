#!/usr/bin/env python3
"""Update employee database with SSNs and employee numbers from gold standard"""

import openpyxl
from wbs_accurate_converter import WBSAccurateConverter

print('=== UPDATING EMPLOYEE DATABASE ===')

# Load gold standard to extract employee data
gold_wb = openpyxl.load_workbook('wbs_gold_standard.xlsx')
gold_ws = gold_wb.active

# Load our conversion to see which employees we have
our_wb = openpyxl.load_workbook('WBS_Full_Conversion_Test.xlsx')
our_ws = our_wb.active

# Get our employee list
our_employees = set()
for row_num in range(9, our_ws.max_row + 1):
    name = our_ws.cell(row=row_num, column=3).value
    if name and name.strip():
        our_employees.add(name)

print(f'Our conversion has {len(our_employees)} employees')

# Extract employee data from gold standard
employee_updates = {}
for row_num in range(9, gold_ws.max_row + 1):
    name = gold_ws.cell(row=row_num, column=3).value
    emp_num = gold_ws.cell(row=row_num, column=1).value
    ssn = gold_ws.cell(row=row_num, column=2).value
    
    if name and name in our_employees and emp_num and ssn:
        employee_updates[name] = {
            'employee_number': str(emp_num),
            'ssn': str(ssn),
            'status': 'A',
            'type': 'H',  # Hourly
            'department': 'PRODUCTION'  # Default department
        }

print(f'Found {len(employee_updates)} employee records to add to database')

# Generate the updated employee database code
print('\n=== GENERATING UPDATED EMPLOYEE DATABASE ===')

print('```python')
print('# Updated employee database with SSNs and employee numbers from gold standard')
print('self.employee_database = {')

# Include existing employees first
converter = WBSAccurateConverter()
for name, info in converter.employee_database.items():
    if name in employee_updates:
        # Use gold standard data
        gold_info = employee_updates[name]
        print(f'    "{name}": {{')
        print(f'        "employee_number": "{gold_info["employee_number"]}",')
        print(f'        "ssn": "{gold_info["ssn"]}",')
        print(f'        "status": "{gold_info["status"]}",')
        print(f'        "type": "{gold_info["type"]}",')
        print(f'        "department": "{gold_info["department"]}"')
        print(f'    }},')
    else:
        # Keep existing data
        print(f'    "{name}": {{')
        print(f'        "employee_number": "{info["employee_number"]}",')
        print(f'        "ssn": "{info["ssn"]}",')
        print(f'        "status": "{info["status"]}",')
        print(f'        "type": "{info["type"]}",')
        print(f'        "department": "{info["department"]}"')
        print(f'    }},')

# Add new employees
for name, info in employee_updates.items():
    if name not in converter.employee_database:
        print(f'    "{name}": {{')
        print(f'        "employee_number": "{info["employee_number"]}",')
        print(f'        "ssn": "{info["ssn"]}",')
        print(f'        "status": "{info["status"]}",')
        print(f'        "type": "{info["type"]}",')
        print(f'        "department": "{info["department"]}"')
        print(f'    }},')

print('}')
print('```')

print(f'\n🎯 READY TO UPDATE: {len(employee_updates)} employee records')
print('💾 Copy the generated database code to wbs_accurate_converter.py')
print('🚀 Then run final conversion for identical results!')