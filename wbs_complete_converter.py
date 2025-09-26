#!/usr/bin/env python3
"""
WBS Complete Sierra Payroll Converter
Complete employee database from actual WBS file with all 71 employees
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import re
from openpyxl import Workbook

class WBSCompleteConverter:
    """
    Converts Sierra payroll to exact WBS format with complete employee database
    """
    
    def __init__(self, gold_master_order_path: Optional[str] = None):
        """Initialize converter with complete employee database"""
        self.employee_database = self._create_employee_database()
        if gold_master_order_path and Path(gold_master_order_path).exists():
            with open(gold_master_order_path, 'r', encoding='utf-8') as f:
                self.gold_master_order = [line.strip() for line in f if line.strip()]
        else:
            self.gold_master_order = list(self.employee_database.keys())
    
    def _create_employee_database(self) -> Dict[str, Dict]:
        """Create complete employee database with exact WBS format data"""
        return {
            "Alcaraz, Luis": {
                "employee_number": "0000659096",
                "ssn": "432946242",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Alvarez, Jose": {
                "employee_number": "0000662584",
                "ssn": "534908967",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Anolin, Robert M": {
                "employee_number": "0000659058",
                "ssn": "552251095",
                "status": "A",
                "type": "S",
                "department": "SALES"
            },
            "Arizmendi, Fernando": {
                "employee_number": "0000659100",
                "ssn": "613871092",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Arroyo, Jose": {
                "employee_number": "0000674796",
                "ssn": "364725751",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Bello, Luis": {
                "employee_number": "0000662081",
                "ssn": "616226754",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Bocanegra, Jose": {
                "employee_number": "0000701064",
                "ssn": "605908531",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Bustos, Eric": {
                "employee_number": "0000700760",
                "ssn": "603965173",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Castaneda, Andy": {
                "employee_number": "0000668812",
                "ssn": "611042001",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Castillo, Moises": {
                "employee_number": "0000675652",
                "ssn": "653246578",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Chavez, Derick J": {
                "employee_number": "0000698157",
                "ssn": "610591002",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Chavez, Endhy": {
                "employee_number": "0000698158",
                "ssn": "625379918",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Contreras, Brian": {
                "employee_number": "0000682812",
                "ssn": "137178003",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Cuevas Barragan, Carlos": {
                "employee_number": "0000681979",
                "ssn": "615873427",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Cuevas, Marcelo": {
                "employee_number": "0000659113",
                "ssn": "625928562",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Dean, Jacob P": {
                "employee_number": "0000659051",
                "ssn": "625154423",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Dean, Joe P": {
                "employee_number": "0000659055",
                "ssn": "556534609",
                "status": "A",
                "type": "C",
                "department": "SALES"
            },
            "Duarte, Esau": {
                "employee_number": "0000701059",
                "ssn": "658836473",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Duarte, Kevin": {
                "employee_number": "0000697052",
                "ssn": "654060734",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Espinoza, Jose Federico": {
                "employee_number": "0000659000",
                "ssn": "607794927",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Esquivel, Kleber": {
                "employee_number": "0000659046",
                "ssn": "615292328",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Flores, Saul Daniel L": {
                "employee_number": "0000674802",
                "ssn": "611882540",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Garcia Garcia, Eduardo": {
                "employee_number": "0000659080",
                "ssn": "621364058",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Garcia, Bryan": {
                "employee_number": "0000659075",
                "ssn": "616259654",
                "status": "A",
                "type": "H",
                "department": "GUTTR"
            },
            "Garcia, Miguel A": {
                "employee_number": "0000659112",
                "ssn": "681068099",
                "status": "A",
                "type": "H",
                "department": "GUTTR"
            },
            "Garrido, Raul": {
                "employee_number": "0000658985",
                "ssn": "657554426",
                "status": "A",
                "type": "S",
                "department": "SALES"
            },
            "Gomez, Jose": {
                "employee_number": "0000658982",
                "ssn": "897981424",
                "status": "A",
                "type": "H",
                "department": "SOLAR"
            },
            "Gonzalez, Alejandro": {
                "employee_number": "0000668811",
                "ssn": "341127082",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Gonzalez, Emanuel": {
                "employee_number": "0000668813",
                "ssn": "627736546",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Gonzalez, Miguel": {
                "employee_number": "0000659063",
                "ssn": "623234232",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Hernandez, Diego": {
                "employee_number": "0000702974",
                "ssn": "652143527",
                "status": "A",
                "type": "H",
                "department": "GUTTR"
            },
            "Hernandez, Sergio": {
                "employee_number": "0000659040",
                "ssn": "618243648",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Lopez, Daniel": {
                "employee_number": "0000659102",
                "ssn": "655411126",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Lopez, Gerwin A": {
                "employee_number": "0000701804",
                "ssn": "189351494",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Lopez, Yair A": {
                "employee_number": "0000659039",
                "ssn": "635455748",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Lopez, Zeferino": {
                "employee_number": "0000659048",
                "ssn": "609226343",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Magallanes, Julio": {
                "employee_number": "0000680753",
                "ssn": "612219002",
                "status": "A",
                "type": "S",
                "department": "SALES"
            },
            "Martinez, Alberto": {
                "employee_number": "0000659009",
                "ssn": "621101210",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Martinez, Emiliano B": {
                "employee_number": "0000659030",
                "ssn": "601903561",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Martinez, Maciel": {
                "employee_number": "0000659038",
                "ssn": "607333861",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Mateos, Daniel": {
                "employee_number": "0000698156",
                "ssn": "660484575",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Moreno, Eduardo": {
                "employee_number": "0000659047",
                "ssn": "610215629",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Nava, Juan M": {
                "employee_number": "0000697056",
                "ssn": "636667958",
                "status": "A",
                "type": "H",
                "department": "SOLAR"
            },
            "Olivares, Alberto M": {
                "employee_number": "0000688269",
                "ssn": "622936952",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Pacheco Estrada, Jesus": {
                "employee_number": "0000675644",
                "ssn": "645935042",
                "status": "A",
                "type": "H",
                "department": "GUTTR"
            },
            "Padilla, Alex": {
                "employee_number": "0000658988",
                "ssn": "569697404",
                "status": "A",
                "type": "S",
                "department": "SALES"
            },
            "Padilla, Carlos": {
                "employee_number": "0000658991",
                "ssn": "614425738",
                "status": "A",
                "type": "H",
                "department": "SOLAR"
            },
            "Pajarito, Ramon": {
                "employee_number": "0000676086",
                "ssn": "685942713",
                "status": "A",
                "type": "H",
                "department": "GUTTR"
            },
            "Pealatere, Francis": {
                "employee_number": "0000675458",
                "ssn": "625098739",
                "status": "A",
                "type": "C",
                "department": "SALES"
            },
            "Pelagio, Miguel Angel": {
                "employee_number": "0000659093",
                "ssn": "086310738",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Perez, Edgar": {
                "employee_number": "0000659101",
                "ssn": "797771646",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Perez, Octavio": {
                "employee_number": "0000698658",
                "ssn": "658873980",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Phein, Saeng Tsing": {
                "employee_number": "0000695183",
                "ssn": "624722627",
                "status": "A",
                "type": "S",
                "department": "SALES"
            },
            "Ramos Grana, Omar": {
                "employee_number": "0000682814",
                "ssn": "645024748",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Rios, Jose D": {
                "employee_number": "0000658996",
                "ssn": "530358447",
                "status": "A",
                "type": "S",
                "department": "SALES"
            },
            "Rivas Beltran, Angel M": {
                "employee_number": "0000665198",
                "ssn": "358119787",
                "status": "A",
                "type": "H",
                "department": "GUTTR"
            },
            "Robledo, Francisco": {
                "employee_number": "0000658979",
                "ssn": "613108074",
                "status": "A",
                "type": "H",
                "department": "SOLAR"
            },
            "Robleza, Dianne": {
                "employee_number": "0000662082",
                "ssn": "626946016",
                "status": "A",
                "type": "H",
                "department": "ADMIN"
            },
            "Rodriguez, Antoni": {
                "employee_number": "0000699565",
                "ssn": "654991245",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Romero Solis, Juan": {
                "employee_number": "0000676689",
                "ssn": "836220003",
                "status": "A",
                "type": "H",
                "department": "GUTTR"
            },
            "Santos, Efrain": {
                "employee_number": "0000659086",
                "ssn": "634473263",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Santos, Javier": {
                "employee_number": "0000659084",
                "ssn": "603297017",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Serrano, Erick V": {
                "employee_number": "0000702972",
                "ssn": "006437019",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Shafer, Emily": {
                "employee_number": "0000659098",
                "ssn": "622809130",
                "status": "A",
                "type": "S",
                "department": "ADMIN"
            },
            "Stokes, Symone": {
                "employee_number": "0000694868",
                "ssn": "616259695",
                "status": "A",
                "type": "H",
                "department": "ADMIN"
            },
            "Torres, Anthony": {
                "employee_number": "0000688270",
                "ssn": "658102450",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Torrez, Jose R": {
                "employee_number": "0000659090",
                "ssn": "625855596",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Valle, Victor": {
                "employee_number": "0000668810",
                "ssn": "602060741",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Vargas Pineda, Karina": {
                "employee_number": "0000702061",
                "ssn": "640670356",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Vera, Victor": {
                "employee_number": "0000659045",
                "ssn": "628795401",
                "status": "A",
                "type": "H",
                "department": "ROOF"
            },
            "Young, Giana L": {
                "employee_number": "0000658972",
                "ssn": "602762103",
                "status": "A",
                "type": "S",
                "department": "ADMIN"
            }
        }
    
    def normalize_name(self, name: str) -> str:
        """Normalize employee name to match WBS format"""
        if not isinstance(name, str) or not name.strip():
            return ""
        
        name = re.sub(r'\s+', ' ', name.strip())
        
        # If already in "Last, First" format, return as is
        if ',' in name:
            return name
        
        # Convert "First Last" to "Last, First"
        parts = name.split()
        if len(parts) >= 2:
            return f"{parts[-1]}, {' '.join(parts[:-1])}"
        
        return name
    
    def find_employee_info(self, name: str) -> Dict:
        """Find employee information in database"""
        normalized_name = self.normalize_name(name)
        
        # Direct match first
        if normalized_name in self.employee_database:
            return self.employee_database[normalized_name]
        
        # Fuzzy matching for variations
        for db_name, info in self.employee_database.items():
            if self._names_match(normalized_name, db_name):
                return info
        
        # Generate default if not found
        return {
            "employee_number": f"UNKNOWN_{hash(normalized_name) % 10000:04d}",
            "ssn": "000000000",
            "status": "A", 
            "type": "H",
            "department": "UNKNOWN"
        }
    
    def _names_match(self, name1: str, name2: str) -> bool:
        """Check if two names refer to the same person"""
        # Remove punctuation and compare
        clean1 = re.sub(r'[^\w\s]', '', name1.lower())
        clean2 = re.sub(r'[^\w\s]', '', name2.lower()) 
        
        # Check if main parts match
        parts1 = clean1.split()
        parts2 = clean2.split()
        
        if len(parts1) >= 2 and len(parts2) >= 2:
            # Check if first and last names match
            return (parts1[0] in parts2 or parts2[0] in parts1) and \
                   (parts1[-1] in parts2 or parts2[-1] in parts1)
        
        return False
    
    def parse_sierra_file(self, file_path: str) -> pd.DataFrame:
        """Parse Sierra Excel file and extract employee time data"""
        try:
            # Read Excel file - Sierra format typically starts from row 1
            df = pd.read_excel(file_path, header=0)
            
            # Clean column names
            df.columns = df.columns.astype(str).str.strip()
            
            # Find employee name column
            name_columns = [col for col in df.columns if 
                          any(keyword in col.lower() for keyword in ['name', 'employee', 'worker'])]
            
            if not name_columns:
                # If no obvious name column, use first text column
                for col in df.columns:
                    if df[col].dtype == 'object':
                        name_columns = [col]
                        break
            
            if not name_columns:
                raise ValueError("Could not identify employee name column")
            
            name_col = name_columns[0]
            
            # Find hours and rate columns
            numeric_columns = [col for col in df.columns if 
                             df[col].dtype in ['int64', 'float64'] and df[col].notna().sum() > 0]
            
            if len(numeric_columns) < 2:
                raise ValueError("Could not identify hours and rate columns")
            
            # Assume first numeric is hours, second is rate (or try to identify by names)
            hours_col = None
            rate_col = None
            
            for col in numeric_columns:
                col_lower = col.lower()
                if 'hour' in col_lower or 'hrs' in col_lower or 'time' in col_lower:
                    hours_col = col
                elif 'rate' in col_lower or 'pay' in col_lower or 'wage' in col_lower:
                    rate_col = col
            
            # If not identified by name, use first two numeric columns
            if hours_col is None:
                hours_col = numeric_columns[0]
            if rate_col is None and len(numeric_columns) > 1:
                rate_col = numeric_columns[1]
            elif rate_col is None:
                rate_col = numeric_columns[0]  # Same as hours if only one numeric column
            
            # Create clean dataframe
            sierra_data = pd.DataFrame({
                'Employee Name': df[name_col],
                'Hours': pd.to_numeric(df[hours_col], errors='coerce'),
                'Rate': pd.to_numeric(df[rate_col], errors='coerce')
            })
            
            # Clean data - remove rows with missing critical data
            sierra_data = sierra_data.dropna(subset=['Employee Name', 'Hours', 'Rate'])
            sierra_data = sierra_data[sierra_data['Hours'] > 0]  # Must have positive hours
            sierra_data = sierra_data[sierra_data['Employee Name'].str.strip() != '']
            
            return sierra_data
            
        except Exception as e:
            print(f"Error parsing Sierra file: {str(e)}")
            raise
    
    def apply_california_overtime_rules(self, hours: float, rate: float) -> Dict[str, float]:
        """Apply California daily overtime rules"""
        regular_hours = 0.0
        ot15_hours = 0.0  # 1.5x overtime
        ot20_hours = 0.0  # 2x overtime
        
        if hours <= 8:
            regular_hours = hours
        elif hours <= 12:
            regular_hours = 8.0
            ot15_hours = hours - 8.0
        else:
            regular_hours = 8.0
            ot15_hours = 4.0  # Hours 8-12
            ot20_hours = hours - 12.0
        
        # Calculate amounts
        regular_amount = regular_hours * rate
        ot15_amount = ot15_hours * rate * 1.5
        ot20_amount = ot20_hours * rate * 2.0
        
        return {
            'regular_hours': regular_hours,
            'ot15_hours': ot15_hours, 
            'ot20_hours': ot20_hours,
            'regular_amount': regular_amount,
            'ot15_amount': ot15_amount,
            'ot20_amount': ot20_amount,
            'total_amount': regular_amount + ot15_amount + ot20_amount
        }
    
    def consolidate_employees(self, sierra_data: pd.DataFrame) -> pd.DataFrame:
        """Consolidate multiple time entries per employee into single records"""
        consolidated = []
        
        # Group by employee name
        for employee_name, group in sierra_data.groupby('Employee Name'):
            normalized_name = self.normalize_name(employee_name)
            
            # Sum all hours for this employee
            total_hours = group['Hours'].sum()
            
            # Use the most common rate (or average if rates vary)
            # In most cases, employees should have consistent rates
            most_common_rate = group['Rate'].mode().iloc[0] if len(group['Rate'].mode()) > 0 else group['Rate'].mean()
            
            consolidated.append({
                'Employee Name': normalized_name,
                'Total Hours': total_hours,
                'Rate': most_common_rate,
                'Record Count': len(group)
            })
        
        return pd.DataFrame(consolidated)

    def create_wbs_excel(self, sierra_data: pd.DataFrame, output_path: str, pre_consolidated: bool = False) -> str:
        """Create WBS format Excel file with exact column structure"""
        
        # Check if data is already consolidated
        if pre_consolidated:
            # Data is already in consolidated format
            consolidated_data = sierra_data
        else:
            # First, consolidate employees
            consolidated_data = self.consolidate_employees(sierra_data)
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "WEEKLY"
        
        # WBS Header structure (EXACT format from actual WBS file)
        headers = [
            ["# V", "DO NOT EDIT", "Version = B90216-00", "FmtRev = 2.1", 
             f"RunTime = {datetime.now().strftime('%Y%m%d-%H%M%S')}", "CliUnqId = 055269", 
             "CliName = Sierra Roofing and Solar Inc", "Freq = W", 
             f"PEDate = {datetime.now().strftime('%m/%d/%Y')}", 
             f"RptDate = {datetime.now().strftime('%m/%d/%Y')}", 
             f"CkDate = {datetime.now().strftime('%m/%d/%Y')}", "EmpType = SSN", 
             "DoNotes = 1", "PayRates = H+;S+;E+;C+", "RateCol = 6", "T1 = 7+", 
             "CodeBeg = 8", "CodeEnd = 26", "NoteCol = 27"] + [None] * 9,
            ["# U", "CliUnqID", "055269"] + [None] * 25,
            ["# N", "Client", "Sierra Roofing and Solar Inc"] + [None] * 25,
            ["# P", "Period End", datetime.now().strftime('%m/%d/%Y')] + [None] * 25,
            ["# R", "Report Due", datetime.now().strftime('%m/%d/%Y')] + [None] * 25,
            ["# C", "Check Date", datetime.now().strftime('%m/%d/%Y')] + [None] * 25,
            ["# B:8", "", "", "", "Pay", "", "", "REGULAR", "OVERTIME", "DOUBLETIME", 
             "VACATION", "SICK", "HOLIDAY", "BONUS", "COMMISSION", "PC HRS MON", "PC TTL MON", 
             "PC HRS TUE", "PC TTL TUE", "PC HRS WED", "PC TTL WED", "PC HRS THU", "PC TTL THU", 
             "PC HRS FRI", "PC TTL FRI", "TRAVEL AMOUNT", "Notes and", None],
            ["# E:26", "SSN", "Employee Name", "Status", "Type", "Pay Rate", "Dept", "A01", "A02", "A03", 
             "A06", "A07", "A08", "A04", "A05", "AH1", "AI1", "AH2", "AI2", "AH3", "AI3", "AH4", "AI4", 
             "AH5", "AI5", "ATE", "Comments", "Totals"]
        ]
        
        # Write headers
        for row_idx, header_row in enumerate(headers, 1):
            for col_idx, value in enumerate(header_row, 1):
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Process consolidated employee data
        current_row = 9  # Start after headers
        
        for _, row in consolidated_data.iterrows():
            employee_name = row['Employee Name']
            total_hours = float(row['Total Hours'])
            rate = float(row['Rate'])
            
            # Get employee info
            emp_info = self.find_employee_info(employee_name)
            
            # Apply California overtime rules to TOTAL hours
            pay_calc = self.apply_california_overtime_rules(total_hours, rate)
            
            # Generate Excel formula for total (like actual WBS format)
            row_num = current_row
            formula = f"=(F{row_num}*H{row_num})+(F{row_num}*I{row_num})+(F{row_num}*J{row_num})+(F{row_num}*K{row_num})+(F{row_num}*L{row_num})+Q{row_num}+S{row_num}+U{row_num}+W{row_num}+Y{row_num}+Z{row_num}"
            
            # WBS Row data (EXACT format matching actual WBS file)
            # CRITICAL: A01-A03 are HOURS, not dollars! Totals are FORMULAS!
            wbs_row = [
                emp_info['employee_number'],    # Col 1: Employee Number
                emp_info['ssn'],                # Col 2: SSN  
                employee_name,                  # Col 3: Employee Name
                emp_info['status'],             # Col 4: Status (A)
                emp_info['type'],               # Col 5: Type (H/S/E/C)
                rate,                           # Col 6: Pay Rate
                emp_info['department'],         # Col 7: Department 
                pay_calc['regular_hours'] if pay_calc['regular_hours'] > 0 else None,  # Col 8: A01 - Regular HOURS
                pay_calc['ot15_hours'] if pay_calc['ot15_hours'] > 0 else None,        # Col 9: A02 - Overtime HOURS  
                pay_calc['ot20_hours'] if pay_calc['ot20_hours'] > 0 else None,        # Col 10: A03 - Doubletime HOURS
                None,                          # Col 11: A06 - Vacation
                None,                          # Col 12: A07 - Sick
                None,                          # Col 13: A08 - Holiday
                None,                          # Col 14: A04 - Bonus
                None,                          # Col 15: A05 - Commission
                None,                          # Col 16: AH1 - PC HRS MON
                None,                          # Col 17: AI1 - PC TTL MON
                None,                          # Col 18: AH2 - PC HRS TUE
                None,                          # Col 19: AI2 - PC TTL TUE
                None,                          # Col 20: AH3 - PC HRS WED
                None,                          # Col 21: AI3 - PC TTL WED
                None,                          # Col 22: AH4 - PC HRS THU
                None,                          # Col 23: AI4 - PC TTL THU
                None,                          # Col 24: AH5 - PC HRS FRI
                None,                          # Col 25: AI5 - PC TTL FRI
                None,                          # Col 26: ATE - Total Extension
                None,                          # Col 27: Comments
                formula                        # Col 28: Excel Formula (like actual WBS)
            ]
            
            # Write row to Excel
            for col_idx, value in enumerate(wbs_row, 1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def convert_sierra_to_wbs(self, input_path: str, output_path: str) -> str:
        """Convert Sierra file to WBS format"""
        try:
            # Parse Sierra file
            sierra_data = self.parse_sierra_file(input_path)
            
            if sierra_data.empty:
                raise ValueError("No valid employee data found in Sierra file")
            
            # Create WBS Excel file
            result_path = self.create_wbs_excel(sierra_data, output_path)
            
            return result_path
            
        except Exception as e:
            print(f"Conversion failed: {str(e)}")
            raise
    
    def convert(self, input_path: str, output_path: str) -> Dict:
        """Main conversion method matching improved_converter interface"""
        try:
            result_path = self.convert_sierra_to_wbs(input_path, output_path)
            return {
                'success': True,
                'output_path': result_path,
                'message': f'Sierra payroll successfully converted to WBS format: {result_path}'
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'message': f'Conversion failed: {str(e)}'
            }

if __name__ == '__main__':
    # Test the converter
    converter = WBSCompleteConverter()
    print(f"Loaded employee database with {len(converter.employee_database)} employees")
    
    # Test conversion
    result = converter.convert('sierra_input_actual.xlsx', 'complete_wbs_test.xlsx')
    print(f"Conversion result: {result}")