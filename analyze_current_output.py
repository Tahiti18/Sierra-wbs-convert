#!/usr/bin/env python3
"""
Analyze current broken output to see what's wrong
"""
import pandas as pd
from openpyxl import load_workbook

def analyze_current_output():
    print("=== CURRENT OUTPUT ANALYSIS ===")
    
    try:
        # Check if current output exists
        wb = load_workbook('/home/user/webapp/wbs_output_current.xlsx')
        ws = wb.active
        
        print(f"Current output - Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # Look for Dianne in current output
        print("\n=== LOOKING FOR DIANNE IN CURRENT OUTPUT ===")
        for row_num in range(1, min(20, ws.max_row + 1)):
            name_cell = ws.cell(row=row_num, column=3).value
            if name_cell and "Dianne" in str(name_cell):
                print(f"Found Dianne in row {row_num}: {name_cell}")
                
                print(f"\n=== DIANNE'S CURRENT DATA (Row {row_num}) ===")
                for col in range(1, min(30, ws.max_column + 1)):
                    cell = ws.cell(row=row_num, column=col)
                    cell_value = cell.value
                    
                    # Check if it's a formula
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        print(f"Col {col}: FORMULA = {cell_value}")
                    else:
                        print(f"Col {col}: {cell_value}")
                break
                        
        # Also check the Sierra input to understand what we're converting
        print("\n=== SIERRA INPUT ANALYSIS ===")
        sierra_df = pd.read_excel('/home/user/webapp/sierra_input_new.xlsx')
        print(f"Sierra shape: {sierra_df.shape}")
        print("Sierra columns:")
        for col in sierra_df.columns:
            print(f"  {col}")
            
        # Find Dianne in Sierra
        dianne_sierra = sierra_df[sierra_df['Employee Name'].str.contains("Dianne", na=False)]
        if not dianne_sierra.empty:
            print("\n=== DIANNE IN SIERRA INPUT ===")
            dianne_row = dianne_sierra.iloc[0]
            for col, value in dianne_row.items():
                print(f"{col}: {value}")
        
    except Exception as e:
        print(f"Error analyzing current output: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_current_output()