#!/usr/bin/env python3
"""Analyze the actual WBS file format to understand the exact requirements"""

import openpyxl
from openpyxl.utils import get_column_letter

print('=== ANALYZING YOUR ACTUAL WBS FILE FORMAT ===')

# Load your actual WBS file
try:
    wb = openpyxl.load_workbook('your_actual_wbs.xlsx')
    ws = wb.active
    print(f'✅ Your WBS file loaded: {ws.max_row} rows, {ws.max_column} columns')
except Exception as e:
    print(f'❌ Could not load WBS file: {e}')
    exit(1)

print('\n=== HEADER STRUCTURE ANALYSIS ===')
for row in range(1, min(10, ws.max_row + 1)):
    print(f'Row {row}:')
    for col in range(1, min(10, ws.max_column + 1)):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            print(f'  Col {get_column_letter(col)}: "{val}"')
    print()

print('\n=== EMPLOYEE DATA ANALYSIS (First 10 employees) ===')

# Find where employee data starts (usually row 9)
data_start_row = 9
employee_count = 0

for row in range(data_start_row, min(data_start_row + 10, ws.max_row + 1)):
    name = ws.cell(row=row, column=3).value  # Column C is employee name
    if name and str(name).strip():
        employee_count += 1
        
        print(f'\n--- EMPLOYEE {employee_count}: {name} (Row {row}) ---')
        
        # Key columns to analyze
        key_cols = [
            (1, 'A', 'Employee Number'),
            (2, 'B', 'SSN'),
            (3, 'C', 'Name'), 
            (4, 'D', 'Status'),
            (5, 'E', 'Type'),
            (6, 'F', 'Rate'),
            (7, 'G', 'Dept'),
            (8, 'H', 'A01'),
            (9, 'I', 'A02'), 
            (10, 'J', 'A03'),
            (11, 'K', 'A06'),
            (12, 'L', 'A07'),
            (13, 'M', 'A08'),
            (28, 'AB', 'Total')
        ]
        
        for col_num, col_letter, col_name in key_cols:
            if col_num <= ws.max_column:
                val = ws.cell(row=row, column=col_num).value
                print(f'  {col_name:15} (Col {col_letter}): {val}')

print(f'\n=== PATTERN ANALYSIS ===')
print(f'Total employees found: {employee_count}')

# Look for patterns in SSNs and Employee Numbers
print('\n=== SSN AND EMPLOYEE NUMBER PATTERNS ===')
ssn_samples = []
emp_num_samples = []

for row in range(data_start_row, min(data_start_row + 5, ws.max_row + 1)):
    name = ws.cell(row=row, column=3).value
    if name and str(name).strip():
        ssn = ws.cell(row=row, column=2).value
        emp_num = ws.cell(row=row, column=1).value
        
        if ssn: ssn_samples.append(str(ssn))
        if emp_num: emp_num_samples.append(str(emp_num))

print(f'SSN samples: {ssn_samples[:3]}')
print(f'Employee number samples: {emp_num_samples[:3]}')

# Check for empty cells vs zeros
print('\n=== EMPTY CELL vs ZERO ANALYSIS ===')
sample_row = data_start_row
if ws.cell(row=sample_row, column=3).value:
    for col in range(11, 27):  # Columns K through Z
        val = ws.cell(row=sample_row, column=col).value
        if val is None:
            print(f'Col {get_column_letter(col)}: EMPTY (None)')
        elif val == 0:
            print(f'Col {get_column_letter(col)}: ZERO (0)')
        else:
            print(f'Col {get_column_letter(col)}: VALUE ({val})')

print('\n=== DEPARTMENT ANALYSIS ===')
dept_samples = []
for row in range(data_start_row, min(data_start_row + 5, ws.max_row + 1)):
    name = ws.cell(row=row, column=3).value
    if name and str(name).strip():
        dept = ws.cell(row=row, column=7).value
        if dept: dept_samples.append(str(dept))

print(f'Department samples: {set(dept_samples)}')

print('\n=== CRITICAL FINDINGS FOR FIXING OUR CONVERTER ===')
print('🎯 Issues to fix based on your actual WBS format:')
print('1. SSN format and source')
print('2. Employee number format and source') 
print('3. Department mapping')
print('4. Empty cells vs zeros in unused columns')
print('5. Header metadata format')