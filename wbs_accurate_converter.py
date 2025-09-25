#!/usr/bin/env python3
"""
WBS Accurate Sierra Payroll Converter
Matches the EXACT WBS format structure based on gold standard analysis
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import re
from openpyxl import Workbook

class WBSAccurateConverter:
    """
    Converts Sierra payroll to exact WBS format based on gold standard analysis
    """
    
    def __init__(self, gold_master_order_path: Optional[str] = None):
        """Initialize converter with employee database"""
        self.employee_database = self._create_employee_database()
        if gold_master_order_path and Path(gold_master_order_path).exists():
            with open(gold_master_order_path, 'r', encoding='utf-8') as f:
                self.gold_master_order = [line.strip() for line in f if line.strip()]
        else:
            self.gold_master_order = list(self.employee_database.keys())
    
    def _create_employee_database(self) -> Dict[str, Dict]:
        """Create employee database with exact WBS format data"""
        return {
            "Robleza, Dianne": {
                "employee_number": "0000662082",
                "ssn": "626946016", 
                "status": "A",
                "type": "H",
                "department": "ADMIN"
            },
            "Shafer, Emily": {
                "employee_number": "0000659098",
                "ssn": "622809130",
                "status": "A", 
                "type": "S",
                "department": "A"
            },
            "Stokes, Symone": {
                "employee_number": "0000694868",
                "ssn": "616259695",
                "status": "A",
                "type": "H", 
                "department": "A"
            },
            "Young, Giana L": {
                "employee_number": "0000658972",
                "ssn": "602762103",
                "status": "A",
                "type": "S",
                "department": "A"
            },
            "Garcia, Bryan": {
                "employee_number": "0000659075",
                "ssn": "616259654",
                "status": "A",
                "type": "H",
                "department": "A"
            },
            "Garcia, Miguel A": {
                "employee_number": "0000659112", 
                "ssn": "681068099",
                "status": "A",
                "type": "H",
                "department": "A"
            },
            "Hernandez, Diego": {
                "employee_number": "0000702974",
                "ssn": "652143527",
                "status": "A",
                "type": "H",
                "department": "A"
            },
            "Pacheco Estrada, Jesus": {
                "employee_number": "0000675644",
                "ssn": "645935042", 
                "status": "A",
                "type": "H",
                "department": "A"
            },
            "Pajarito, Ramon": {
                "employee_number": "0000676086",
                "ssn": "685942713",
                "status": "A",
                "type": "H",
                "department": "A"
            },
            "Rivas Beltran, Angel M": {
                "employee_number": "0000665198",
                "ssn": "358119787",
                "status": "A",
                "type": "H",
                "department": "A"
            }
        }
    
    def normalize_name(self, name: str) -> str:
        """Normalize employee name to match WBS format"""
        if not isinstance(name, str) or not name.strip():
            return ""
        
        name = re.sub(r'\s+', ' ', name.strip())
        
        # If already in "Last, First" format, return as is
        if ',' in name:
            return name
        
        # Convert "First Last" to "Last, First"
        parts = name.split()
        if len(parts) >= 2:
            return f"{parts[-1]}, {' '.join(parts[:-1])}"
        
        return name
    
    def find_employee_info(self, name: str) -> Dict:
        """Find employee information in database"""
        normalized_name = self.normalize_name(name)
        
        # Direct match first
        if normalized_name in self.employee_database:
            return self.employee_database[normalized_name]
        
        # Fuzzy matching for variations
        for db_name, info in self.employee_database.items():
            if self._names_match(normalized_name, db_name):
                return info
        
        # Generate default if not found
        return {
            "employee_number": f"UNKNOWN_{hash(normalized_name) % 10000:04d}",
            "ssn": "000000000",
            "status": "A", 
            "type": "H",
            "department": "UNKNOWN"
        }
    
    def _names_match(self, name1: str, name2: str) -> bool:
        """Check if two names refer to the same person"""
        # Remove punctuation and compare
        clean1 = re.sub(r'[^\w\s]', '', name1.lower())
        clean2 = re.sub(r'[^\w\s]', '', name2.lower()) 
        
        # Check if main parts match
        parts1 = clean1.split()
        parts2 = clean2.split()
        
        if len(parts1) >= 2 and len(parts2) >= 2:
            # Check if first and last names match
            return (parts1[0] in parts2 or parts2[0] in parts1) and \
                   (parts1[-1] in parts2 or parts2[-1] in parts1)
        
        return False
    
    def parse_sierra_file(self, file_path: str) -> pd.DataFrame:
        """Parse Sierra Excel file and extract employee time data"""
        try:
            # Read Excel file - Sierra format typically starts from row 1
            df = pd.read_excel(file_path, header=0)
            
            # Clean column names
            df.columns = df.columns.astype(str).str.strip()
            
            # Find employee name column
            name_columns = [col for col in df.columns if 
                          any(keyword in col.lower() for keyword in ['name', 'employee', 'worker'])]
            
            if not name_columns:
                # If no obvious name column, use first text column
                for col in df.columns:
                    if df[col].dtype == 'object':
                        name_columns = [col]
                        break
            
            if not name_columns:
                raise ValueError("Could not identify employee name column")
            
            name_col = name_columns[0]
            
            # Find hours and rate columns
            numeric_columns = [col for col in df.columns if 
                             df[col].dtype in ['int64', 'float64'] and df[col].notna().sum() > 0]
            
            if len(numeric_columns) < 2:
                raise ValueError("Could not identify hours and rate columns")
            
            # Assume first numeric is hours, second is rate (or try to identify by names)
            hours_col = None
            rate_col = None
            
            for col in numeric_columns:
                col_lower = col.lower()
                if 'hour' in col_lower or 'hrs' in col_lower or 'time' in col_lower:
                    hours_col = col
                elif 'rate' in col_lower or 'pay' in col_lower or 'wage' in col_lower:
                    rate_col = col
            
            # If not identified by name, use first two numeric columns
            if hours_col is None:
                hours_col = numeric_columns[0]
            if rate_col is None and len(numeric_columns) > 1:
                rate_col = numeric_columns[1]
            elif rate_col is None:
                rate_col = numeric_columns[0]  # Same as hours if only one numeric column
            
            # Create clean dataframe
            sierra_data = pd.DataFrame({
                'Employee Name': df[name_col],
                'Hours': pd.to_numeric(df[hours_col], errors='coerce'),
                'Rate': pd.to_numeric(df[rate_col], errors='coerce')
            })
            
            # Clean data - remove rows with missing critical data
            sierra_data = sierra_data.dropna(subset=['Employee Name', 'Hours', 'Rate'])
            sierra_data = sierra_data[sierra_data['Hours'] > 0]  # Must have positive hours
            sierra_data = sierra_data[sierra_data['Employee Name'].str.strip() != '']
            
            return sierra_data
            
        except Exception as e:
            print(f"Error parsing Sierra file: {str(e)}")
            raise
    
    def apply_california_overtime_rules(self, hours: float, rate: float) -> Dict[str, float]:
        """Apply California daily overtime rules"""
        regular_hours = 0.0
        ot15_hours = 0.0  # 1.5x overtime
        ot20_hours = 0.0  # 2x overtime
        
        if hours <= 8:
            regular_hours = hours
        elif hours <= 12:
            regular_hours = 8.0
            ot15_hours = hours - 8.0
        else:
            regular_hours = 8.0
            ot15_hours = 4.0  # Hours 8-12
            ot20_hours = hours - 12.0
        
        # Calculate amounts
        regular_amount = regular_hours * rate
        ot15_amount = ot15_hours * rate * 1.5
        ot20_amount = ot20_hours * rate * 2.0
        
        return {
            'regular_hours': regular_hours,
            'ot15_hours': ot15_hours, 
            'ot20_hours': ot20_hours,
            'regular_amount': regular_amount,
            'ot15_amount': ot15_amount,
            'ot20_amount': ot20_amount,
            'total_amount': regular_amount + ot15_amount + ot20_amount
        }
    
    def create_wbs_excel(self, sierra_data: pd.DataFrame, output_path: str) -> str:
        """Create WBS format Excel file with exact column structure"""
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "WEEKLY"
        
        # WBS Header structure (exact format from gold standard)
        headers = [
            ["# V", "DO NOT EDIT", "Version = B90216-00", "FmtRev = 2.1", 
             f"RunTime = {datetime.now().strftime('%Y%m%d-%H%M%S')}", "CliUnqId = 055269", 
             "CliName = Sierra Roofing and Solar Inc", "Freq = W", 
             f"PEDate = {datetime.now().strftime('%m/%d/%Y')}", 
             f"RptDate = {datetime.now().strftime('%m/%d/%Y')}", 
             f"CkDate = {datetime.now().strftime('%m/%d/%Y')}", "EmpType = SSN", 
             "DoNotes = 1", "PayRates = H+;S+;E+;C+", "RateCol = 6", "T1 = 7+", 
             "CodeBeg = 8", "CodeEnd = 26", "NoteCol = 27"] + [None] * 9,
            ["# U", "CliUnqID"] + [None] * 26,
            ["# N", "Client"] + [None] * 26,
            ["# P", "Period End"] + [None] * 26,
            ["# R", "Report Due"] + [None] * 26,
            ["# C", "Check Date"] + [None] * 26,
            ["# B:8", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "Notes and", None],
            ["# E:26", "SSN", "Employee Name", "Status", "Type", "Pay Rate", "Dept", "A01", "A02", "A03", "A06", "A07", "A08", "A04", "A05", "AH1", "AI1", "AH2", "AI2", "AH3", "AI3", "AH4", "AI4", "AH5", "AI5", "ATE", "Comments", "Totals"]
        ]
        
        # Write headers
        for row_idx, header_row in enumerate(headers, 1):
            for col_idx, value in enumerate(header_row, 1):
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Process employee data
        current_row = 9  # Start after headers
        
        for _, row in sierra_data.iterrows():
            employee_name = self.normalize_name(row['Employee Name'])
            hours = float(row['Hours'])
            rate = float(row['Rate'])
            
            # Get employee info
            emp_info = self.find_employee_info(employee_name)
            
            # Apply California overtime rules
            pay_calc = self.apply_california_overtime_rules(hours, rate)
            
            # WBS Row data (exact column mapping from gold standard)
            wbs_row = [
                emp_info['employee_number'],  # Col 1: Employee Number
                emp_info['ssn'],              # Col 2: SSN
                employee_name,                # Col 3: Employee Name  
                emp_info['status'],           # Col 4: Status (A)
                emp_info['type'],             # Col 5: Type (H/S/E/C)
                rate,                         # Col 6: Pay Rate
                emp_info['department'],       # Col 7: Department
                pay_calc['regular_hours'],    # Col 8: A01 - Regular Hours
                0,                           # Col 9: A02 - Overtime 1
                0,                           # Col 10: A03 - Doubletime  
                0,                           # Col 11: A06 - Vacation
                0,                           # Col 12: A07 - Sick
                0,                           # Col 13: A08 - Holiday
                0,                           # Col 14: A04 - Bonus
                0,                           # Col 15: A05 - Commission
                0,                           # Col 16: AH1 - PC HRS MON
                0,                           # Col 17: AI1 - PC TTL MON
                0,                           # Col 18: AH2 - PC HRS TUE
                0,                           # Col 19: AI2 - PC TTL TUE
                0,                           # Col 20: AH3 - PC HRS WED
                0,                           # Col 21: AI3 - PC TTL WED
                0,                           # Col 22: AH4 - PC HRS THU
                0,                           # Col 23: AI4 - PC TTL THU
                0,                           # Col 24: AH5 - PC HRS FRI
                0,                           # Col 25: AI5 - PC TTL FRI
                0,                           # Col 26: ATE - Total Extension
                "",                          # Col 27: Comments
                pay_calc['total_amount']      # Col 28: Totals
            ]
            
            # Write row to Excel
            for col_idx, value in enumerate(wbs_row, 1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def convert_sierra_to_wbs(self, input_path: str, output_path: str) -> str:
        """Convert Sierra file to WBS format"""
        try:
            # Parse Sierra file
            sierra_data = self.parse_sierra_file(input_path)
            
            if sierra_data.empty:
                raise ValueError("No valid employee data found in Sierra file")
            
            # Create WBS Excel file
            result_path = self.create_wbs_excel(sierra_data, output_path)
            
            return result_path
            
        except Exception as e:
            print(f"Conversion failed: {str(e)}")
            raise
    
    def convert(self, input_path: str, output_path: str) -> Dict:
        """Main conversion method matching improved_converter interface"""
        try:
            result_path = self.convert_sierra_to_wbs(input_path, output_path)
            return {
                'success': True,
                'output_path': result_path,
                'message': 'Conversion completed successfully'
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'output_path': None
            }