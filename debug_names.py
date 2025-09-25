#!/usr/bin/env python3
"""
Debug name formats in fixed output
"""
import pandas as pd
from openpyxl import load_workbook

def debug_names():
    print("=== DEBUGGING NAME FORMATS ===")
    
    # Load fixed output
    wb = load_workbook('/home/user/webapp/wbs_output_FIXED.xlsx')
    ws = wb.active
    
    print("First 20 rows, Column 3 (Names):")
    for row in range(1, 21):
        name = ws.cell(row=row, column=3).value
        print(f"Row {row:2d}: {name}")
        
        if name and "ianne" in str(name).lower():
            print(f"    *** FOUND DIANNE VARIANT IN ROW {row}! ***")

if __name__ == "__main__":
    debug_names()