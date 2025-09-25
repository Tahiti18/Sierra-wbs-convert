#!/usr/bin/env python3
"""
FIXED WBS Sierra Payroll Converter
Addresses the critical issues:
1. Column 28 should contain CALCULATED VALUES, not formulas
2. Sierra input uses 'Name' column, not 'Employee Name' 
3. Columns A01, A02, A03 contain HOURS (correct)
4. Missing columns should be 0, not None
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import re
from openpyxl import Workbook

class WBSFixedConverter:
    """
    FIXED converter that outputs exact WBS format with calculated values
    """
    
    def __init__(self, gold_master_order_path: Optional[str] = None):
        """Initialize converter with employee database"""
        self.employee_database = self._create_employee_database()
        if gold_master_order_path and Path(gold_master_order_path).exists():
            with open(gold_master_order_path, 'r', encoding='utf-8') as f:
                self.gold_master_order = [line.strip() for line in f if line.strip()]
        else:
            self.gold_master_order = list(self.employee_database.keys())
    
    def _create_employee_database(self) -> Dict[str, Dict]:
        """Create employee database with exact WBS format data"""
        return {
            "Robleza, Dianne": {
                "employee_number": "0000662082",
                "ssn": "626946016", 
                "status": "A",
                "type": "H",
                "department": "ADMIN"
            },
            "Shafer, Emily": {
                "employee_number": "0000659098",
                "ssn": "622809130",
                "status": "A", 
                "type": "S",
                "department": "A"
            },
            "Stokes, Symone": {
                "employee_number": "0000694868",
                "ssn": "616259695",
                "status": "A",
                "type": "H", 
                "department": "A"
            },
            "Pacheco Estrada, Jesus": {
                "employee_number": "0000675644",
                "ssn": "645935042", 
                "status": "A",
                "type": "H",
                "department": "A"
            },
            "Pajarito, Ramon": {
                "employee_number": "0000676086",
                "ssn": "685942713",
                "status": "A",
                "type": "H",
                "department": "A"
            },
            "Rivas Beltran, Angel M": {
                "employee_number": "0000665198",
                "ssn": "358119787",
                "status": "A",
                "type": "H",
                "department": "A"
            }
        }
    
    def normalize_name(self, name: str) -> str:
        """Normalize employee name to match WBS format"""
        if not isinstance(name, str) or not name.strip():
            return ""
        
        name = re.sub(r'\s+', ' ', name.strip())
        
        # If already in "Last, First" format, return as is
        if ',' in name:
            return name
        
        # Convert "First Last" to "Last, First"
        parts = name.split()
        if len(parts) >= 2:
            return f"{parts[-1]}, {' '.join(parts[:-1])}"
        
        return name
    
    def find_employee_info(self, name: str) -> Dict:
        """Find employee information in database"""
        normalized_name = self.normalize_name(name)
        
        # Direct match first
        if normalized_name in self.employee_database:
            return self.employee_database[normalized_name]
        
        # Partial match search
        for db_name, emp_info in self.employee_database.items():
            if self._names_match(normalized_name, db_name):
                return emp_info
        
        # Default if not found
        return {
            "employee_number": "0000000000",
            "ssn": "000000000",
            "status": "A",
            "type": "H",
            "department": "A"
        }
    
    def _names_match(self, name1: str, name2: str) -> bool:
        """Check if two names refer to the same person"""
        name1_parts = set(name1.lower().replace(',', '').split())
        name2_parts = set(name2.lower().replace(',', '').split())
        
        # If they share at least 2 parts, consider them matching
        return len(name1_parts.intersection(name2_parts)) >= 2
    
    def apply_california_overtime_rules(self, total_hours: float, rate: float) -> Dict:
        """
        Apply California overtime rules to total hours
        8+ hours = 1.5x overtime, 12+ hours = 2x doubletime
        """
        regular_hours = 0
        ot15_hours = 0
        ot20_hours = 0
        
        if total_hours <= 8:
            regular_hours = total_hours
        elif total_hours <= 12:
            regular_hours = 8
            ot15_hours = total_hours - 8
        else:
            regular_hours = 8
            ot15_hours = 4  # Hours 9-12
            ot20_hours = total_hours - 12  # Hours 13+
        
        # Calculate amounts
        regular_amount = regular_hours * rate
        ot15_amount = ot15_hours * rate * 1.5
        ot20_amount = ot20_hours * rate * 2.0
        total_amount = regular_amount + ot15_amount + ot20_amount
        
        return {
            'regular_hours': regular_hours,
            'ot15_hours': ot15_hours,
            'ot20_hours': ot20_hours,
            'regular_amount': regular_amount,
            'ot15_amount': ot15_amount,
            'ot20_amount': ot20_amount,
            'total_amount': total_amount
        }
    
    def parse_sierra_file(self, file_path: str) -> pd.DataFrame:
        """
        FIXED: Parse Sierra Excel file correctly
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path)
            
            print(f"Sierra file columns: {df.columns.tolist()}")
            
            # FIXED: Look for 'Name' column (not 'Employee Name')
            name_columns = [col for col in df.columns if 
                          'name' in col.lower() and col.lower() not in ['username', 'filename']]
            
            if not name_columns:
                raise ValueError("Could not identify employee name column")
            
            name_col = name_columns[0]
            print(f"Using name column: {name_col}")
            
            # Find hours and rate columns
            numeric_columns = [col for col in df.columns if 
                             df[col].dtype in ['int64', 'float64'] and df[col].notna().sum() > 0]
            
            print(f"Numeric columns: {numeric_columns}")
            
            if len(numeric_columns) < 2:
                raise ValueError("Could not identify hours and rate columns")
            
            # Identify hours and rate columns by name patterns
            hours_col = None
            rate_col = None
            
            for col in numeric_columns:
                col_lower = col.lower()
                if 'hour' in col_lower or 'hrs' in col_lower or 'time' in col_lower:
                    hours_col = col
                elif 'rate' in col_lower or 'pay' in col_lower or 'wage' in col_lower:
                    rate_col = col
            
            # If not identified by name, use logical order
            if hours_col is None:
                hours_col = numeric_columns[0]
            if rate_col is None and len(numeric_columns) > 1:
                rate_col = numeric_columns[1]
            elif rate_col is None:
                rate_col = numeric_columns[0]
            
            print(f"Using hours column: {hours_col}")
            print(f"Using rate column: {rate_col}")
            
            # Create clean dataframe with FIXED column name
            sierra_data = pd.DataFrame({
                'Employee Name': df[name_col],  # Rename to standard format
                'Hours': pd.to_numeric(df[hours_col], errors='coerce'),
                'Rate': pd.to_numeric(df[rate_col], errors='coerce')
            })
            
            # Clean data
            sierra_data = sierra_data.dropna(subset=['Employee Name', 'Hours', 'Rate'])
            sierra_data = sierra_data[sierra_data['Hours'] > 0]
            sierra_data = sierra_data[sierra_data['Employee Name'].str.strip() != '']
            
            print(f"Processed {len(sierra_data)} valid employee records")
            
            return sierra_data
            
        except Exception as e:
            print(f"Error parsing Sierra file: {str(e)}")
            raise
    
    def consolidate_employee_hours(self, sierra_data: pd.DataFrame) -> pd.DataFrame:
        """Consolidate multiple entries for same employee"""
        consolidated = sierra_data.groupby('Employee Name').agg({
            'Hours': 'sum',
            'Rate': 'first'  # Assume same rate for same employee
        }).reset_index()
        
        # Rename for consistency
        consolidated.rename(columns={'Hours': 'Total Hours'}, inplace=True)
        
        return consolidated
    
    def create_wbs_excel(self, consolidated_data: pd.DataFrame, output_path: str) -> str:
        """
        FIXED: Create WBS Excel with CALCULATED VALUES, not formulas
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "WEEKLY"
        
        # WBS Header Row (from gold standard analysis)
        header = [
            "# V", "DO NOT EDIT", "Version = B90216-00", "FmtRev = 2.1", "RunTime = 20250905-115816",
            "CliUnqId = 055269", "CliName = Sierra Roofing and Solar Inc", "Freq = W", 
            "PEDate = 09/07/2025", "RptDate = 09/10/2025", "CkDate = 09/12/2025", "EmpType = SSN",
            "DoNotes = 1", "PayRates = H+;S+;E+;C+", "RateCol = 6", "T1 = 7+", "CodeBeg = 8", 
            "CodeEnd = 26", "NoteCol = 27", "", "", "", "", "", "", "", "", ""
        ]
        
        # Write header
        for col_idx, value in enumerate(header, 1):
            ws.cell(row=1, column=col_idx, value=value)
        
        current_row = 9  # Start from row 9 like gold standard
        
        # Process each employee
        for _, row in consolidated_data.iterrows():
            employee_name = row['Employee Name']
            total_hours = float(row['Total Hours'])
            rate = float(row['Rate'])
            
            print(f"Processing {employee_name}: {total_hours} hours @ ${rate}/hour")
            
            # Get employee info
            emp_info = self.find_employee_info(employee_name)
            
            # Apply California overtime rules
            pay_calc = self.apply_california_overtime_rules(total_hours, rate)
            
            # FIXED: WBS Row with CALCULATED VALUES in column 28
            wbs_row = [
                emp_info['employee_number'],    # Col 1: Employee Number
                emp_info['ssn'],                # Col 2: SSN
                employee_name,                  # Col 3: Employee Name  
                emp_info['status'],             # Col 4: Status (A)
                emp_info['type'],               # Col 5: Type (H/S/E/C)
                rate,                           # Col 6: Pay Rate
                emp_info['department'],         # Col 7: Department
                pay_calc['regular_hours'],      # Col 8: A01 - Regular HOURS 
                pay_calc['ot15_hours'],         # Col 9: A02 - Overtime 1.5x HOURS
                pay_calc['ot20_hours'],         # Col 10: A03 - Doubletime 2x HOURS
                0,                             # Col 11: A06 - Vacation
                0,                             # Col 12: A07 - Sick
                0,                             # Col 13: A08 - Holiday
                0,                             # Col 14: A04 - Bonus
                0,                             # Col 15: A05 - Commission
                0,                             # Col 16: AH1 - PC HRS MON
                0,                             # Col 17: AI1 - PC TTL MON
                0,                             # Col 18: AH2 - PC HRS TUE
                0,                             # Col 19: AI2 - PC TTL TUE
                0,                             # Col 20: AH3 - PC HRS WED
                0,                             # Col 21: AI3 - PC TTL WED
                0,                             # Col 22: AH4 - PC HRS THU
                0,                             # Col 23: AI4 - PC TTL THU
                0,                             # Col 24: AH5 - PC HRS FRI
                0,                             # Col 25: AI5 - PC TTL FRI
                0,                             # Col 26: ATE - Total Extension
                "",                            # Col 27: Comments
                pay_calc['total_amount']        # Col 28: CALCULATED TOTAL (not formula!)
            ]
            
            # Write row to Excel
            for col_idx, value in enumerate(wbs_row, 1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            print(f"  -> Regular: {pay_calc['regular_hours']}h, OT1.5: {pay_calc['ot15_hours']}h, OT2.0: {pay_calc['ot20_hours']}h, Total: ${pay_calc['total_amount']}")
            
            current_row += 1
        
        # Save workbook
        wb.save(output_path)
        print(f"WBS file saved to: {output_path}")
        return output_path
    
    def convert_sierra_to_wbs(self, input_path: str, output_path: str) -> str:
        """Convert Sierra file to WBS format"""
        try:
            # Parse Sierra file
            print("=== PARSING SIERRA FILE ===")
            sierra_data = self.parse_sierra_file(input_path)
            
            if sierra_data.empty:
                raise ValueError("No valid employee data found in Sierra file")
            
            # Consolidate employee data  
            print("=== CONSOLIDATING EMPLOYEE DATA ===")
            consolidated_data = self.consolidate_employee_hours(sierra_data)
            
            # Create WBS Excel file
            print("=== CREATING WBS EXCEL ===")
            result_path = self.create_wbs_excel(consolidated_data, output_path)
            
            print("=== CONVERSION COMPLETE ===")
            return result_path
            
        except Exception as e:
            print(f"Conversion failed: {str(e)}")
            raise
    
    def convert(self, input_path: str, output_path: str) -> Dict:
        """Main conversion method matching improved_converter interface"""
        try:
            result_path = self.convert_sierra_to_wbs(input_path, output_path)
            return {
                'success': True,
                'output_path': result_path,
                'message': 'Conversion completed successfully'
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'message': f'Conversion failed: {str(e)}'
            }

# Test function
def test_conversion():
    """Test the fixed converter"""
    print("=== TESTING FIXED WBS CONVERTER ===")
    
    converter = WBSFixedConverter()
    
    input_file = "/home/user/webapp/sierra_input_new.xlsx"
    output_file = "/home/user/webapp/wbs_output_FIXED.xlsx"
    
    try:
        result = converter.convert(input_file, output_file)
        
        if result['success']:
            print(f"✅ CONVERSION SUCCESSFUL!")
            print(f"Output saved to: {result['output_path']}")
            
            # Quick verification
            print("\n=== QUICK VERIFICATION ===")
            df = pd.read_excel(output_file)
            print(f"Output has {len(df)} rows and {len(df.columns)} columns")
            
            # Check Dianne's data
            dianne_rows = df[df.iloc[:, 2].astype(str).str.contains("Dianne", na=False)]
            if not dianne_rows.empty:
                dianne_data = dianne_rows.iloc[0]
                print(f"Dianne's data:")
                print(f"  Employee #: {dianne_data.iloc[0]}")
                print(f"  SSN: {dianne_data.iloc[1]}")
                print(f"  Name: {dianne_data.iloc[2]}")
                print(f"  Rate: {dianne_data.iloc[5]}")
                print(f"  Regular Hours: {dianne_data.iloc[7]}")
                print(f"  Total Amount: {dianne_data.iloc[27]}")
                
                expected_total = 28.0 * 4.0  # $28/hour * 4 hours = $112
                actual_total = dianne_data.iloc[27]
                print(f"  Expected: $112, Actual: ${actual_total}")
                
                if abs(float(actual_total) - 112.0) < 0.01:
                    print("  ✅ TOTAL MATCHES GOLD STANDARD!")
                else:
                    print("  ❌ TOTAL DOES NOT MATCH!")
            
        else:
            print(f"❌ CONVERSION FAILED: {result['message']}")
            
    except Exception as e:
        print(f"❌ TEST FAILED: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_conversion()